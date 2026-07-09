"""US-186 — Organigramas con utilización por scope.

Cubre:
- TC-186-1: matriz mensual a nivel programa suma participaciones por
  recurso a través de los proyectos del programa (y excluye proyectos
  fuera del programa), con %FTE total tenant aparte.
- TC-186-2: export de programa (XLSX 2 hojas) con fill de alerta rojo
  en el mes >100% y amarillo en ≥80%.
- TC-186-3: export de organización y global (tenant) responden 200 con
  attachment; el global incluye recursos de todos los proyectos.
- TC-186-4: export de proyecto conserva las 4 hojas US-150 y agrega
  "Recursos (FTE)" + "Uso mensual".
"""
from datetime import date, timedelta
from io import BytesIO

import pytest
from openpyxl import load_workbook
from sqlalchemy import select

from app.models.tenant import Tenant
from app.services.capacity import monthly_utilization
from tests.factories import create_admin_role, create_tenant, create_user, login


async def _setup(client, db_session):
    t = await create_tenant(db_session)
    tid = str(t.id)
    admin_role = await create_admin_role(db_session, t)
    await create_user(
        db_session, tenant=t, username="admin", email="admin@acme.example.com",
        password="Str0ng-Admin-1!", roles=[admin_role],
    )
    auth = await login(client, "admin", "Str0ng-Admin-1!")
    r = await client.post("/api/v1/organizations", json={"name": "Org1"}, headers=auth["_authz"])
    org_id = r.json()["id"]
    pr = await client.post(
        "/api/v1/programs",
        json={"name": "Programa X", "organization_id": org_id},
        headers=auth["_authz"],
    )
    assert pr.status_code == 201, pr.text
    program_id = pr.json()["id"]
    me = await client.get("/api/v1/auth/me", headers=auth["_authz"])
    pm_id = me.json()["id"]

    async def mk_project(name, in_program=True):
        body = {
            "name": name, "description": "US-186", "type": "transformation",
            "priority": 3, "organization_id": org_id, "pm_id": pm_id,
        }
        if in_program:
            body["program_id"] = program_id
        p = await client.post("/api/v1/projects", json=body, headers=auth["_authz"])
        assert p.status_code == 201, p.text
        return p.json()["id"]

    p1 = await mk_project("Prog Uno")
    p2 = await mk_project("Prog Dos")
    p3 = await mk_project("Fuera", in_program=False)
    tenant = (
        await db_session.execute(select(Tenant).where(Tenant.id == tid))
    ).scalar_one()
    return auth, tenant, org_id, program_id, (p1, p2, p3)


async def _actor_and_assign(client, auth, projects_pcts, name="Carlos Mejia"):
    r = await client.post(
        "/api/v1/actors",
        json={"name": name, "project_capacity_pct": 80, "portfolio_function": "arquitectura"},
        headers=auth["_authz"],
    )
    actor_id = r.json()["id"]
    today = date.today()
    for pid, pct in projects_pcts:
        rr = await client.post(
            f"/api/v1/projects/{pid}/participations",
            json={
                "actor_id": actor_id, "allocation_pct": pct,
                "start_date": today.replace(day=1).isoformat(),
                "end_date": (today + timedelta(days=120)).isoformat(),
            },
            headers=auth["_authz"],
        )
        assert rr.status_code == 201, rr.text
    return actor_id


@pytest.mark.asyncio
async def test_us186_program_matrix_sums_by_resource(client, db_session):
    auth, tenant, _, program_id, (p1, p2, p3) = await _setup(client, db_session)
    a = await _actor_and_assign(client, auth, [(p1, 60), (p2, 45), (p3, 30)])

    util = await monthly_utilization(
        db_session, tenant, scope_type="program", scope_id=program_id
    )
    row = next(r for r in util["rows"] if r["actor_id"] == a)
    # Scope programa: 60+45 = 105 (el proyecto fuera NO suma al scope).
    assert row["per_month"][0] == 105
    assert row["scope_current_pct"] == 105
    # Total tenant sí incluye el proyecto fuera: 135.
    assert row["tenant_current_pct"] == 135
    assert row["projects_count"] == 2
    assert row["alert_months"] >= 1
    assert len(util["months"]) == 12


@pytest.mark.asyncio
async def test_us186_program_export_with_alert_fills(client, db_session):
    auth, _, _, program_id, (p1, p2, _) = await _setup(client, db_session)
    await _actor_and_assign(client, auth, [(p1, 60), (p2, 45)], name="Rojo Mensual")
    await _actor_and_assign(client, auth, [(p1, 85)], name="Amarillo Mensual")

    r = await client.get(
        f"/api/v1/programs/{program_id}/organigrama/export", headers=auth["_authz"]
    )
    assert r.status_code == 200, r.text
    assert "attachment" in r.headers["content-disposition"]
    wb = load_workbook(BytesIO(r.content))
    assert wb.sheetnames == ["Organigrama", "Uso mensual"]
    ws = wb["Uso mensual"]
    fills = {}
    for row in ws.iter_rows(min_row=2, max_row=3):
        name = row[0].value
        cell = row[1]  # primer mes
        fills[name] = (cell.value, cell.fill.start_color.rgb)
    # 105 > 100 → rojo (FEE2E2); 85 ≥ 80 → amarillo (FEF3C7).
    assert fills["Rojo Mensual"][0] == 105
    assert str(fills["Rojo Mensual"][1]).endswith("FEE2E2")
    assert fills["Amarillo Mensual"][0] == 85
    assert str(fills["Amarillo Mensual"][1]).endswith("FEF3C7")
    # Columna final "Meses en alerta".
    header = [c.value for c in ws[1]]
    assert header[-1] == "Meses en alerta"


@pytest.mark.asyncio
async def test_us186_org_and_tenant_exports(client, db_session):
    auth, _, org_id, _, (p1, _, p3) = await _setup(client, db_session)
    await _actor_and_assign(client, auth, [(p1, 20)], name="Solo Programa")
    await _actor_and_assign(client, auth, [(p3, 20)], name="Solo Fuera")

    r = await client.get(
        f"/api/v1/organizations/{org_id}/organigrama/export", headers=auth["_authz"]
    )
    assert r.status_code == 200, r.text
    wb = load_workbook(BytesIO(r.content))
    names = [row[0].value for row in wb["Organigrama"].iter_rows(min_row=2)]
    assert "Solo Programa" in names and "Solo Fuera" in names

    g = await client.get("/api/v1/capacity/organigrama/export", headers=auth["_authz"])
    assert g.status_code == 200, g.text
    wbg = load_workbook(BytesIO(g.content))
    gnames = [row[0].value for row in wbg["Organigrama"].iter_rows(min_row=2)]
    assert "Solo Programa" in gnames and "Solo Fuera" in gnames


@pytest.mark.asyncio
async def test_us186_project_export_gains_fte_sheets(client, db_session):
    auth, _, _, _, (p1, _, _) = await _setup(client, db_session)
    await _actor_and_assign(client, auth, [(p1, 50)], name="Del Proyecto")

    r = await client.get(
        f"/api/v1/projects/{p1}/organigrama/export", headers=auth["_authz"]
    )
    assert r.status_code == 200, r.text
    wb = load_workbook(BytesIO(r.content))
    assert wb.sheetnames == [
        "Áreas", "Equipos", "Roles", "Recursos", "Recursos (FTE)", "Uso mensual"
    ]
    names = [row[0].value for row in wb["Recursos (FTE)"].iter_rows(min_row=2)]
    assert "Del Proyecto" in names
