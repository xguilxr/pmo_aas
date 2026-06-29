"""ENH-152 — RAID export service (unit).

Verifica que el Excel tiene las 4 hojas RAID en español con sus headers, y
que los builders resuelven nombres (área, responsable actor-fallback-user),
el estado en ES y la fecha de creación de negocio. No requiere DB.
"""
from __future__ import annotations

from datetime import UTC, date, datetime
from io import BytesIO
from types import SimpleNamespace

from openpyxl import load_workbook

from app.services.raid_export import (
    AID_HEADERS,
    RISK_HEADERS,
    build_issue_rows,
    build_risk_rows,
    export_raid_xlsx,
    export_single_sheet_xlsx,
)


def _load(data: bytes):
    return load_workbook(BytesIO(data))


def test_export_single_sheet_has_one_sheet_with_headers():
    # ENH-168: export individual por tipo = una sola hoja.
    data = export_single_sheet_xlsx(
        title="Riesgos",
        headers=RISK_HEADERS,
        rows=[["R-1", "Riesgo X", "d", 12, "Mitigando", "PMO", "Ana", date(2026, 1, 5)]],
    )
    wb = _load(data)
    assert wb.sheetnames == ["Riesgos"]
    assert [c.value for c in wb["Riesgos"][1]] == RISK_HEADERS
    assert wb["Riesgos"]["B2"].value == "Riesgo X"


def test_raid_export_has_four_spanish_sheets_with_headers():
    data = export_raid_xlsx(
        risks_rows=[["R-1", "Riesgo X", "desc", 12, "Mitigando", "PMO", "Ana", date(2026, 1, 5)]],
        actions_rows=[["A-1", "Acción Y", "desc", 2, "Abierta", "Ing", "Beto", date(2026, 2, 1)]],
        incidents_rows=[],
        decisions_rows=[],
    )
    wb = _load(data)
    assert wb.sheetnames == ["Riesgos", "Acciones", "Incidencias", "Decisiones"]

    riesgos = wb["Riesgos"]
    assert [c.value for c in riesgos[1]] == RISK_HEADERS
    assert riesgos["A2"].value == "R-1"
    assert riesgos["F2"].value == "PMO"  # Responsable área
    assert riesgos["G2"].value == "Ana"  # Responsable

    acciones = wb["Acciones"]
    assert [c.value for c in acciones[1]] == AID_HEADERS
    assert acciones["B2"].value == "Acción Y"


def test_raid_export_empty_sheets_have_headers_only():
    data = export_raid_xlsx(
        risks_rows=[], actions_rows=[], incidents_rows=[], decisions_rows=[]
    )
    wb = _load(data)
    assert wb.sheetnames == ["Riesgos", "Acciones", "Incidencias", "Decisiones"]
    assert wb["Decisiones"].max_row == 1  # solo header
    assert [c.value for c in wb["Incidencias"][1]] == AID_HEADERS


def test_build_risk_rows_resolves_names_status_and_business_date():
    r = SimpleNamespace(
        folio="R-1", title="Riesgo", description="d", severity=9,
        status="resolved", area_id="area-1",
        owner_actor_id="actor-1", owner_id="user-1",
        identified_at=date(2026, 3, 10),
    )
    rows = build_risk_rows(
        [r],
        area_names={"area-1": "Operaciones"},
        actor_names={"actor-1": "Carla Actor"},
        user_names={"user-1": "Usuario Login"},
    )
    # Actor preferido; status en ES; fecha = identified_at.
    assert rows[0] == [
        "R-1", "Riesgo", "d", 9, "Resuelto",
        "Operaciones", "Carla Actor", date(2026, 3, 10),
    ]


def test_build_issue_rows_actor_fallback_to_user_and_reported_date():
    i = SimpleNamespace(
        folio="A-1", title="Acción", description="", priority=3,
        status="in_progress", area_id="area-1",
        owner_actor_id=None, owner_id="user-1",
        reported_at=datetime(2026, 4, 2, 15, 30, tzinfo=UTC),
    )
    rows = build_issue_rows(
        [i],
        area_names={"area-1": "Ingeniería"},
        actor_names={},
        user_names={"user-1": "Usuario Login"},
    )
    # Sin actor → fallback a Usuario; reported_at → solo fecha; status en ES.
    assert rows[0] == [
        "A-1", "Acción", "", 3, "En Progreso",
        "Ingeniería", "Usuario Login", date(2026, 4, 2),
    ]
