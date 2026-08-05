"""ENH-193 — export/download backend consistente con la plantilla V1.

- PLAN_HEADERS = 15 columnas de la plantilla (round-trip sin mapeo).
- Orden de filas = orden real del plan (WBS natural), no outline-first.
- WBS exportado en formato texto (BUG-088).
- Round-trip: regenerate_xlsx → parse_xlsx re-lee estado/WBS/avance.
"""
from __future__ import annotations

from datetime import date

from openpyxl import load_workbook

from app.models.task import Task
from app.services.plan_regenerator import (
    PLAN_HEADERS,
    PlanExportContext,
    plan_order,
    regenerate_xlsx,
)
from app.services.xlsx_task_parser import parse_xlsx


def _task(**kw) -> Task:
    base = {
        "tenant_id": "t1",
        "project_id": "p1",
        "name": "Tarea",
        "progress": 0,
        "is_milestone": False,
        "status": "not_started",
    }
    base.update(kw)
    return Task(**base)


def test_headers_match_template_v1():
    assert PLAN_HEADERS == [
        "WBS", "Tarea", "Outline Level", "Inicio", "Fin",
        "Duración (días)", "Avance (%)", "Estado", "Área Responsable",
        "Responsable", "Criticidad", "Es hito", "Hito Relacionado",
        "Predecessors", "Successors",
    ]


def test_plan_order_natural_wbs_not_outline_first():
    """Antes: outline-first agrupaba todos los nivel-1 y después los
    nivel-2. Ahora 1 < 1.2 < 1.10 < 2 (jerárquico natural)."""
    tasks = [
        _task(id="a", wbs_code="1.10", outline_level=2),
        _task(id="b", wbs_code="2", outline_level=1),
        _task(id="c", wbs_code="1", outline_level=1),
        _task(id="d", wbs_code="1.2", outline_level=2),
    ]
    assert [t.wbs_code for t in plan_order(tasks)] == ["1", "1.2", "1.10", "2"]


def test_regenerate_xlsx_roundtrip():
    tasks = [
        _task(
            id="m1", wbs_code="1", name="Hito final", is_milestone=True,
            status="in_progress", progress=50, outline_level=1,
            start_date=date(2026, 1, 5), end_date=date(2026, 1, 9),
            duration_days=5, area_id="area-1", assignee_actor_id="act-1",
            is_critical=True, predecessors=["2"],
        ),
        _task(
            id="t2", wbs_code="1.30", name="Sub", status="completed",
            progress=100, outline_level=2, related_milestone_id="m1",
        ),
    ]
    ctx = PlanExportContext(
        area_names={"area-1": "PMO"},
        actor_names={"act-1": "Juan Pérez"},
        milestone_wbs={"m1": "1"},
    )
    data = regenerate_xlsx(tasks, ctx)

    wb = load_workbook(__import__("io").BytesIO(data))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    assert list(rows[0]) == PLAN_HEADERS
    # Fila del hito: área/responsable/criticidad legibles, estado crudo.
    assert rows[1][7] == "in_progress"
    assert rows[1][8] == "PMO"
    assert rows[1][9] == "Juan Pérez"
    assert rows[1][10] == "Sí"
    assert rows[1][11] == "Sí"
    assert rows[1][13] == "2"
    # Fila 1.30: hito relacionado resuelto a WBS; WBS en formato texto.
    assert rows[2][0] == "1.30"
    assert rows[2][12] == "1"
    assert ws.cell(row=3, column=1).number_format == "@"

    # Round-trip: el import re-lee el archivo sin pérdida.
    parsed = parse_xlsx(data)
    assert [t.wbs_code for t in parsed.tasks] == ["1", "1.30"]
    assert [t.status for t in parsed.tasks] == ["in_progress", "completed"]
    assert [t.progress for t in parsed.tasks] == [50, 100]
    assert parsed.tasks[0].resources_raw == "Juan Pérez"
    assert parsed.tasks[0].area_raw == "PMO"
    assert parsed.tasks[1].related_milestone_wbs == "1"
    assert parsed.tasks[0].predecessors_raw == "2"
