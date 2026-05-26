"""US-150 — Organigrama export service (unit).

Verifica que el Excel generado tiene las 4 hojas esperadas con sus
headers, y que las filas de datos se escriben en orden. No requiere DB.
"""
from __future__ import annotations

from io import BytesIO

from openpyxl import load_workbook

from app.services.organigrama_export import export_organigrama_xlsx


def _load(data: bytes):
    return load_workbook(BytesIO(data))


def test_organigrama_export_has_four_sheets_with_headers():
    data = export_organigrama_xlsx(
        areas_rows=[["a1", "PMO", "Oficina de proyectos", "Ana", "Sí"]],
        teams_rows=[["t1", "PMO", "Core", "", "Sí"]],
        roles_rows=[["r1", "PM", "Project Manager", "Sí"]],
        recursos_rows=[
            ["ac1", "Ana", "Core", "PMO", "PM", "ACME", "ana@x.com", "555", "", "Sí"],
        ],
    )
    wb = _load(data)
    assert wb.sheetnames == ["Áreas", "Equipos", "Roles", "Recursos"]

    areas = wb["Áreas"]
    assert [c.value for c in areas[1]] == ["ID", "Nombre", "Descripción", "Líder", "Activa"]
    assert areas["B2"].value == "PMO"

    recursos = wb["Recursos"]
    assert [c.value for c in recursos[1]] == [
        "ID", "Nombre", "Equipo", "Área", "Puesto", "Empresa",
        "Email", "Teléfono", "Manager", "Activo",
    ]
    assert recursos["B2"].value == "Ana"
    assert recursos["G2"].value == "ana@x.com"


def test_organigrama_export_empty_sheets_have_headers_only():
    data = export_organigrama_xlsx(
        areas_rows=[], teams_rows=[], roles_rows=[], recursos_rows=[]
    )
    wb = _load(data)
    assert wb.sheetnames == ["Áreas", "Equipos", "Roles", "Recursos"]
    # Solo header row (sin datos).
    assert wb["Roles"].max_row == 1
    assert [c.value for c in wb["Roles"][1]] == ["ID", "Nombre", "Descripción", "Activo"]
