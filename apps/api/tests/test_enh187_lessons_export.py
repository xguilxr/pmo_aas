"""ENH-187 — Lessons export service (unit).

Sigue el mismo patrón que test_enh186_changes_export.py: verifica que el
Excel tiene 1 hoja "Lecciones" con headers en español, que el builder
resuelve categoría/fase/estado a texto ES, responsable (UUID de Actor →
texto) y tags (lista → string join). No requiere DB.
"""
from __future__ import annotations

from io import BytesIO
from types import SimpleNamespace

from openpyxl import load_workbook

from app.services.lessons_export import (
    LESSON_HEADERS,
    build_lesson_rows,
    export_lessons_xlsx,
)


def _load(data: bytes):
    return load_workbook(BytesIO(data))


def test_export_lessons_has_one_sheet_with_headers():
    data = export_lessons_xlsx(
        rows=[
            [
                "LEC-1",
                "Lección X",
                "Descripción",
                "Éxito",
                "Ejecución",
                "Ana",
                "Recomendación",
                "onboarding, alcance",
                "Publicada",
            ]
        ]
    )
    wb = _load(data)
    assert wb.sheetnames == ["Lecciones"]
    assert [c.value for c in wb["Lecciones"][1]] == LESSON_HEADERS
    assert wb["Lecciones"]["B2"].value == "Lección X"


def test_export_lessons_empty_has_headers_only():
    data = export_lessons_xlsx(rows=[])
    wb = _load(data)
    assert wb.sheetnames == ["Lecciones"]
    assert wb["Lecciones"].max_row == 1
    assert [c.value for c in wb["Lecciones"][1]] == LESSON_HEADERS


def test_build_lesson_rows_resolves_category_phase_owner_and_tags():
    l = SimpleNamespace(
        folio="LEC-1",
        title="Lección de cierre",
        description="Lo que aprendimos",
        category="success",
        phase="closed",
        owner_actor_id="actor-1",
        recommendation="Repetir el patrón",
        tags=["onboarding", "comunicación"],
        status="published",
    )
    rows = build_lesson_rows([l], actor_names={"actor-1": "Ana Responsable"})
    assert rows[0] == [
        "LEC-1",
        "Lección de cierre",
        "Lo que aprendimos",
        "Éxito",
        "Cierre",
        "Ana Responsable",
        "Repetir el patrón",
        ["onboarding", "comunicación"],
        "Publicada",
    ]


def test_build_lesson_rows_handles_missing_owner_category_and_tags():
    l = SimpleNamespace(
        folio="LEC-2",
        title="Lección sin responsable",
        description=None,
        category=None,
        phase=None,
        owner_actor_id=None,
        recommendation=None,
        tags=[],
        status="published",
    )
    rows = build_lesson_rows([l], actor_names={})
    assert rows[0] == [
        "LEC-2",
        "Lección sin responsable",
        "",
        "",
        "",
        "",
        "",
        [],
        "Publicada",
    ]
