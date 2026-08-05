"""ENH-202 — todo lo que el usuario se lleva sale en Helvetica.

US-193 puso Helvetica en el Excel del plan y dejó escrito que era «el primer
cambio masivo de fuente». Esto es el resto: XLSX del backend, PDF, DOCX y el
Excel del portafolio salían cada uno con la suya.

Estas pruebas **abren el archivo generado** y miran qué fuente quedó dentro, que
es lo único que demuestra el requisito. Comprobar que el código llama al helper
mide el llamado, no el resultado — y en este cambio esa diferencia importó dos
veces:

- La primera versión de `aplicar_a_workbook` asignaba un `Font` nuevo al estilo
  `Normal`. openpyxl lo añadía como fuente número 1 y el `cellXf` por defecto
  seguía apuntando a la 0: el archivo salía con Helvetica declarada y **ninguna
  celda usándola**.
- La segunda mutaba `wb._fonts[0].name`. Eso sí funciona, y **contamina**: ese
  `Font` es el mismo objeto en todos los libros del proceso, así que cambiaba
  la fuente de cualquier Excel generado después. Y hacía pasar esta misma
  prueba por el motivo equivocado, que es la peor forma de tenerla en verde.

Lo que NO se comprueba aquí: que el PDF acabe renderizando Helvetica. Eso
depende de que la imagen instale `fonts-urw-base35`, y se vigila leyendo el
`Dockerfile` — construirla ataría la suite a un demonio Docker.
"""
from __future__ import annotations

import io
import re
import zipfile
from pathlib import Path

import pytest

from app.core.tipografia import FUENTE, PILA_CSS, aplicar_a_docx, aplicar_a_workbook

RAIZ_API = Path(__file__).resolve().parents[1]

#: Los sitios que declaran tipografía para algo que el usuario se lleva.
DECLARAN_FUENTE = [
    "app/templates/pdf/base.html",
    "app/services/html_report_renderer.py",
    "app/api/v1/endpoints/reports.py",
    "app/services/reports/gantt_renderer.py",
]


# ---------------------------------------------------------------------------
# XLSX — se abre el archivo y se mira a qué fuente apuntan las celdas
# ---------------------------------------------------------------------------


def _estilos_del_xlsx(wb) -> str:
    buf = io.BytesIO()
    wb.save(buf)
    with zipfile.ZipFile(io.BytesIO(buf.getvalue())) as z:
        return z.read("xl/styles.xml").decode()


def test_las_celdas_sin_estilo_propio_salen_en_helvetica():
    """El caso corriente: una celda de datos, sin `Font` ninguna."""
    from openpyxl import Workbook

    wb = Workbook()
    aplicar_a_workbook(wb)
    wb.active["A1"] = "una celda cualquiera"

    estilos = _estilos_del_xlsx(wb)
    fuentes = re.findall(r'<name val="([^"]+)"/>', estilos)
    id_por_defecto = int(re.search(r'<cellXfs.*?<xf [^>]*fontId="(\d+)"', estilos, re.S).group(1))

    assert fuentes[id_por_defecto] == FUENTE, (
        f"Las celdas sin estilo apuntan a «{fuentes[id_por_defecto]}». Declarar "
        f"Helvetica sin que ninguna celda la use es el defecto que ENH-202 "
        f"corrige, no el arreglo."
    )


def test_no_se_cuela_calibri():
    """openpyxl arranca con Calibri; si sobrevive, el archivo sale mezclado."""
    from openpyxl import Workbook

    wb = Workbook()
    aplicar_a_workbook(wb)
    wb.active["A1"] = "x"

    assert "Calibri" not in _estilos_del_xlsx(wb)


@pytest.mark.parametrize(
    "modulo",
    ["raid_export", "lessons_export", "change_export", "organigrama_export"],
)
def test_los_exports_del_backend_aplican_la_fuente(modulo):
    """Un `Workbook()` nuevo que no pase por el helper vuelve a Calibri."""
    fuente = (RAIZ_API / "app" / "services" / f"{modulo}.py").read_text(encoding="utf-8")

    creaciones = fuente.count("wb = Workbook()")
    aplicaciones = fuente.count("aplicar_a_workbook(wb)")

    assert creaciones and aplicaciones == creaciones, (
        f"{modulo}.py crea {creaciones} libros y aplica la fuente {aplicaciones} "
        f"veces (ENH-202)."
    )


# ---------------------------------------------------------------------------
# DOCX
# ---------------------------------------------------------------------------


def test_el_docx_lleva_helvetica_en_el_estilo_normal():
    from docx import Document

    doc = Document()
    aplicar_a_docx(doc)
    doc.add_paragraph("hola")

    buf = io.BytesIO()
    doc.save(buf)
    with zipfile.ZipFile(io.BytesIO(buf.getvalue())) as z:
        estilos = z.read("word/styles.xml").decode()
    normal = re.search(r'w:styleId="Normal".*?</w:style>', estilos, re.S).group(0)
    declaradas = re.findall(r'w:(?:ascii|hAnsi|eastAsia)="([^"]+)"', normal)

    assert declaradas.count(FUENTE) >= 3, (
        f"El estilo Normal declara {declaradas}. Falta alguna de "
        f"ascii/hAnsi/eastAsia — sin `eastAsia`, Word mezcla dos tipografías."
    )


@pytest.mark.parametrize(
    "ruta", ["app/services/charter_generator.py", "app/services/minutes_formatter.py"]
)
def test_los_docx_del_producto_aplican_la_fuente(ruta):
    assert "aplicar_a_docx(doc)" in (RAIZ_API / ruta).read_text(encoding="utf-8")


# ---------------------------------------------------------------------------
# PDF / SVG — lo que se declara, y lo que hace falta para que sea verdad
# ---------------------------------------------------------------------------


@pytest.mark.parametrize("ruta", DECLARAN_FUENTE)
def test_ninguna_plantilla_pide_una_fuente_que_no_esta_en_la_imagen(ruta):
    """DM Sans es la fuente de marca de la web y **no** está en la imagen.

    Pedirla en un PDF no la trae: fontconfig cae a la sans genérica y el
    informe sale con una tipografía que nadie eligió.
    """
    texto = (RAIZ_API / ruta).read_text(encoding="utf-8")
    lineas = [
        n
        for n, linea in enumerate(texto.splitlines(), 1)
        if "DM Sans" in linea and not linea.lstrip().startswith(("#", "//", "*", "<!--"))
    ]

    assert not lineas, f"{ruta} sigue pidiendo DM Sans en las líneas {lineas}"


@pytest.mark.parametrize("ruta", DECLARAN_FUENTE)
def test_helvetica_va_primera(ruta):
    texto = (RAIZ_API / ruta).read_text(encoding="utf-8")
    familias = re.findall(r"font-family\s*[:=]\s*[\"']?([^;\"'\n}]+)", texto)

    assert familias, f"{ruta} ya no declara ninguna familia"
    for familia in familias:
        primera = familia.split(",")[0].strip().strip("\"'}").strip()
        # `monospace` e `inherit` son familias genéricas deliberadas: código
        # embebido y tablas que heredan del cuerpo. No son una fuente de marca.
        assert primera.lower() in {"helvetica", "monospace", "inherit"}, (
            f"{ruta} declara «{primera}» antes que Helvetica (ENH-202)."
        )


def test_la_imagen_instala_una_fuente_con_metricas_de_helvetica():
    """Sin el paquete, `font-family: Helvetica` es una declaración vacía.

    Medido con WeasyPrint 68.1: con solo `fonts-dejavu-core`, `fc-match
    Helvetica` devuelve `DejaVuSans.ttf`; con `fonts-urw-base35`, el PDF
    incrusta `NimbusSans`.
    """
    dockerfile = (RAIZ_API / "Dockerfile").read_text(encoding="utf-8")
    instalados = [
        linea.strip().rstrip("\\").strip()
        for linea in dockerfile.splitlines()
        if not linea.lstrip().startswith("#")
    ]

    assert "fonts-urw-base35" in instalados, (
        "El Dockerfile no instala una fuente con métricas de Helvetica. "
        "Sin ella, todo lo demás de ENH-202 en los PDF es decorativo."
    )


def test_no_quedan_tipografias_remotas():
    """AM-12: generar un PDF no debe depender de que Google responda.

    El modelo de amenazas pedía «empotrar las tipografías. Se cruza con
    ENH-202», y con Helvetica ya en la imagen el enlace remoto sobraba.
    """
    culpables = []
    for ruta in RAIZ_API.rglob("app/**/*.py"):
        for n, linea in enumerate(ruta.read_text(encoding="utf-8").splitlines(), 1):
            if ("fonts.googleapis.com" in linea or "fonts.gstatic.com" in linea) and not linea.lstrip().startswith("#"):
                culpables.append(f"{ruta.relative_to(RAIZ_API)}:{n}")
    for ruta in RAIZ_API.rglob("app/templates/**/*.html"):
        if "fonts.googleapis.com" in ruta.read_text(encoding="utf-8"):
            culpables.append(str(ruta.relative_to(RAIZ_API)))

    assert not culpables, f"Siguen pidiendo tipografías remotas: {culpables}"


def test_la_pila_css_nombra_la_fuente_que_de_verdad_hay():
    """`Nimbus Sans` explícita: si se cae el alias de fontconfig, no se nota."""
    assert PILA_CSS.startswith("Helvetica")
    assert "Nimbus Sans" in PILA_CSS
