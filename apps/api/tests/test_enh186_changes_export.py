"""ENH-186 — Change Requests export service (unit).

Sigue el mismo patrón que test_enh152_raid_export.py: verifica que el
Excel tiene 1 hoja "Cambios" con headers en español, que el builder
resuelve tipo/estado a texto ES, nombres de solicitante/aprobador
(UUID → texto) y fechas de negocio (requested_at/approved_at). No
requiere DB.
"""
from __future__ import annotations

from datetime import UTC, datetime
from io import BytesIO
from types import SimpleNamespace

from openpyxl import load_workbook

from app.services.change_export import (
    CHANGE_HEADERS,
    build_change_rows,
    export_changes_xlsx,
)


def _load(data: bytes):
    return load_workbook(BytesIO(data))


def test_export_changes_has_one_sheet_with_headers():
    data = export_changes_xlsx(
        rows=[["CHG-1", "Cambio X", "Alcance", "Aprobado", "Ana", "2026-01-05", "Beto", "2026-01-10", "impacto"]]
    )
    wb = _load(data)
    assert wb.sheetnames == ["Cambios"]
    assert [c.value for c in wb["Cambios"][1]] == CHANGE_HEADERS
    assert wb["Cambios"]["B2"].value == "Cambio X"


def test_export_changes_empty_has_headers_only():
    data = export_changes_xlsx(rows=[])
    wb = _load(data)
    assert wb.sheetnames == ["Cambios"]
    assert wb["Cambios"].max_row == 1
    assert [c.value for c in wb["Cambios"][1]] == CHANGE_HEADERS


def test_build_change_rows_resolves_type_status_and_names():
    c = SimpleNamespace(
        folio="CHG-1",
        title="Cambio de alcance",
        type="scope",
        status="approved",
        requested_by="user-1",
        requested_at=datetime(2026, 1, 5, 10, 0, tzinfo=UTC),
        approved_by="user-2",
        approved_at=datetime(2026, 1, 10, 9, 0, tzinfo=UTC),
        impact="Impacto esperado",
    )
    rows = build_change_rows(
        [c],
        user_names={"user-1": "Ana Solicitante", "user-2": "Beto Aprobador"},
    )
    assert rows[0] == [
        "CHG-1",
        "Cambio de alcance",
        "Alcance",
        "Aprobado",
        "Ana Solicitante",
        datetime(2026, 1, 5).date(),
        "Beto Aprobador",
        datetime(2026, 1, 10).date(),
        "Impacto esperado",
    ]


def test_build_change_rows_handles_missing_approver_and_impact():
    c = SimpleNamespace(
        folio="CHG-2",
        title="Cambio en revisión",
        type="cost",
        status="in_review",
        requested_by="user-1",
        requested_at=datetime(2026, 2, 1, 8, 0, tzinfo=UTC),
        approved_by=None,
        approved_at=None,
        impact=None,
    )
    rows = build_change_rows([c], user_names={"user-1": "Ana Solicitante"})
    assert rows[0] == [
        "CHG-2",
        "Cambio en revisión",
        "Costo",
        "En revisión",
        "Ana Solicitante",
        datetime(2026, 2, 1).date(),
        "",
        "",
        "",
    ]
