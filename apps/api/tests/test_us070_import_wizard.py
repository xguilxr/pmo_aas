"""US-070 — Wizard de mapeo de columnas Excel/CSV/MPP.

Cubre los endpoints nuevos `POST /tasks/import/preview` y
`POST /tasks/import/{job_id}/confirm`. El backend real usa Redis
(via `rate_limit.py` y el nuevo `import_job_store.py`); para tests
reemplazamos el cliente redis con un stub in-memory que respeta
el TTL aproximadamente.
"""
from __future__ import annotations

import io
import time
from typing import Any

import pytest
from openpyxl import Workbook

from app.services import import_job_store
from tests.factories import create_admin_role, create_tenant, create_user, login


class _FakeRedis:
    """Stub minimal para `redis.Redis` — solo lo que usa `import_job_store`.

    Guarda `{key: (value, expires_at_monotonic)}`. `set(ex=)` fija TTL;
    `get` devuelve None si expiró; `delete` remueve.
    """

    def __init__(self):
        self._store: dict[str, tuple[str, float]] = {}

    def set(self, key: str, value: str, ex: int | None = None) -> None:
        expiry = time.monotonic() + ex if ex else float("inf")
        self._store[key] = (value, expiry)

    def get(self, key: str) -> str | None:
        row = self._store.get(key)
        if row is None:
            return None
        value, expiry = row
        if time.monotonic() > expiry:
            del self._store[key]
            return None
        return value

    def delete(self, key: str) -> int:
        return int(bool(self._store.pop(key, None)))

    def expire_now(self, key: str) -> None:
        """Helper de test: fuerza expiración inmediata sin esperar."""
        if key in self._store:
            value, _ = self._store[key]
            self._store[key] = (value, 0)


@pytest.fixture(autouse=True)
def _stub_redis(monkeypatch):
    fake = _FakeRedis()
    monkeypatch.setattr(import_job_store, "_get_client", lambda: fake)
    yield fake


# --- builders de fixtures de archivo ---


def _build_xlsx(rows: list[list[Any]], sheet_name: str = "Plan") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_xlsx_multi_sheets(
    sheets: dict[str, list[list[Any]]],
) -> bytes:
    wb = Workbook()
    # Workbook crea una sheet default — la removemos para empezar limpio.
    wb.remove(wb.active)
    for name, rows in sheets.items():
        ws = wb.create_sheet(title=name)
        for row in rows:
            ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


SIMPLE_HEADERS = ["Nombre", "WBS", "Inicio", "Fin", "Duración", "% completado"]
SIMPLE_ROWS = [
    SIMPLE_HEADERS,
    ["Analisis", "1", "2026-01-01", "2026-01-10", 10, 50],
    ["Desarrollo", "2", "2026-01-11", "2026-02-10", 31, 10],
    ["Go-Live", "3", "2026-02-11", "2026-02-11", 0, 0],
]


async def _setup(client, db_session):
    t = await create_tenant(db_session)
    admin_role = await create_admin_role(db_session, t)
    await create_user(
        db_session,
        tenant=t,
        username="admin",
        email="admin@acme.example.com",
        password="Str0ng-Admin-1!",
        roles=[admin_role],
    )
    auth = await login(client, "admin", "Str0ng-Admin-1!")
    org = await client.post(
        "/api/v1/organizations", json={"name": "Org1"}, headers=auth["_authz"]
    )
    me = await client.get("/api/v1/auth/me", headers=auth["_authz"])
    p = await client.post(
        "/api/v1/projects",
        json={
            "name": "PMSP",
            "description": "d",
            "type": "bau",
            "priority": 3,
            "organization_id": org.json()["id"],
            "pm_id": me.json()["id"],
        },
        headers=auth["_authz"],
    )
    return auth, p.json()["id"]


# --- TC-070.1 — Excel multi-sheet: preview devuelve los nombres ---


@pytest.mark.asyncio
async def test_tc070_1_xlsx_multi_sheet_preview(client, db_session):
    auth, proj_id = await _setup(client, db_session)
    data = _build_xlsx_multi_sheets(
        {
            "Plan Q1": SIMPLE_ROWS,
            "Plan Q2": [SIMPLE_HEADERS, ["Otra tarea", "1", None, None, 5, 0]],
            "Notas": [["texto suelto"]],
        }
    )
    files = {
        "file": (
            "plan.xlsx",
            data,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    }
    r = await client.post(
        f"/api/v1/projects/{proj_id}/tasks/import/preview",
        files=files,
        headers=auth["_authz"],
    )
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["source"] == "xlsx"
    assert body["sheets"] == ["Plan Q1", "Plan Q2", "Notas"]
    # Sin `sheet` en query, usa la primera (Plan Q1).
    assert body["sheet_used"] == "Plan Q1"
    assert body["task_count"] == 3

    # Preview con sheet explícito.
    files = {
        "file": (
            "plan.xlsx",
            data,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    }
    r2 = await client.post(
        f"/api/v1/projects/{proj_id}/tasks/import/preview?sheet=Plan%20Q2",
        files=files,
        headers=auth["_authz"],
    )
    assert r2.status_code == 200
    assert r2.json()["sheet_used"] == "Plan Q2"
    assert r2.json()["task_count"] == 1


# --- TC-070.2 — CSV con delimitador ; y BOM UTF-8 ---


@pytest.mark.asyncio
async def test_tc070_2_csv_semicolon_bom(client, db_session):
    auth, proj_id = await _setup(client, db_session)
    csv_text = (
        "﻿Nombre;WBS;Inicio;Fin;Duración;% completado\n"
        "Kickoff;1;2026-01-01;2026-01-05;5;20\n"
        "Diseño;2;2026-01-06;2026-02-01;20;0\n"
    )
    data = csv_text.encode("utf-8")
    files = {"file": ("plan.csv", data, "text/csv")}
    r = await client.post(
        f"/api/v1/projects/{proj_id}/tasks/import/preview",
        files=files,
        headers=auth["_authz"],
    )
    assert r.status_code == 200, r.text
    assert r.json()["source"] == "csv"
    assert r.json()["task_count"] == 2
    assert r.json()["sheets"] == []  # CSV no tiene sheets


# --- TC-070.3 — Mapping manual override ---


@pytest.mark.asyncio
async def test_tc070_3_manual_mapping_override(client, db_session):
    """Si el archivo tiene headers no reconocidos, el usuario puede
    enviar un mapping manual en el confirm.
    """
    auth, proj_id = await _setup(client, db_session)
    # Headers custom que no matchean HEADER_ALIASES — el parser auto-detect
    # falla si el `name` no se reconoce. El mapping manual lo arregla.
    data = _build_xlsx(
        [
            ["Actividad", "Código", "Desde", "Hasta", "Días"],
            ["Task A", "1.1", "2026-01-01", "2026-01-05", 5],
            ["Task B", "1.2", "2026-01-06", "2026-01-10", 5],
        ]
    )
    files = {
        "file": (
            "custom.xlsx",
            data,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    }
    r_prev = await client.post(
        f"/api/v1/projects/{proj_id}/tasks/import/preview",
        files=files,
        headers=auth["_authz"],
    )
    # Los headers "Actividad" + "Código" NO están en HEADER_ALIASES →
    # columns_detected vacío → task_count=0 pero preview se guarda.
    assert r_prev.status_code == 200, r_prev.text
    body = r_prev.json()
    assert body["task_count"] == 0
    job_id = body["job_id"]

    # Confirm con mapping explícito apuntando "name" a col 0, "wbs_code" a col 1.
    r_conf = await client.post(
        f"/api/v1/projects/{proj_id}/tasks/import/{job_id}/confirm",
        json={
            "mapping": {
                "name": 0,
                "wbs_code": 1,
                "start_date": 2,
                "end_date": 3,
                "duration_days": 4,
            },
            "strategy": "replace",
        },
        headers=auth["_authz"],
    )
    assert r_conf.status_code == 200, r_conf.text
    assert r_conf.json()["imported"] == 2


# --- TC-070.4 — Mapping sin name → 422 ---


@pytest.mark.asyncio
async def test_tc070_4_confirm_missing_name_422(client, db_session):
    auth, proj_id = await _setup(client, db_session)
    data = _build_xlsx(SIMPLE_ROWS)
    files = {
        "file": (
            "plan.xlsx",
            data,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    }
    r_prev = await client.post(
        f"/api/v1/projects/{proj_id}/tasks/import/preview",
        files=files,
        headers=auth["_authz"],
    )
    job_id = r_prev.json()["job_id"]

    r_conf = await client.post(
        f"/api/v1/projects/{proj_id}/tasks/import/{job_id}/confirm",
        json={"mapping": {"wbs_code": 1}, "strategy": "replace"},
        headers=auth["_authz"],
    )
    assert r_conf.status_code == 422
    assert "name" in r_conf.text.lower()


# --- TC-070.5 — job_id expirado → 410 ---


@pytest.mark.asyncio
async def test_tc070_5_expired_job_id_410(client, db_session, _stub_redis):
    auth, proj_id = await _setup(client, db_session)
    data = _build_xlsx(SIMPLE_ROWS)
    files = {
        "file": (
            "plan.xlsx",
            data,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    }
    r_prev = await client.post(
        f"/api/v1/projects/{proj_id}/tasks/import/preview",
        files=files,
        headers=auth["_authz"],
    )
    job_id = r_prev.json()["job_id"]

    # Forzar expiración del job.
    _stub_redis.expire_now(f"import:job:{job_id}")

    r_conf = await client.post(
        f"/api/v1/projects/{proj_id}/tasks/import/{job_id}/confirm",
        json={"strategy": "replace"},
        headers=auth["_authz"],
    )
    assert r_conf.status_code == 410
    assert "expir" in r_conf.text.lower()


# --- TC-070.6 — MPP también pasa por preview/confirm ---


@pytest.mark.asyncio
async def test_tc070_6_mpp_wizard_flow(client, db_session, monkeypatch):
    """El wizard acepta .mpp igual que XLSX/CSV, pero el mapping no
    aplica (MPXJ ya emite shape normalizado). Mockeamos el parser MPP
    que invoca java via subprocess.
    """
    import json
    import shutil
    import subprocess

    from app.services.msproject import mpp_parser

    monkeypatch.setattr(shutil, "which", lambda _: "/usr/bin/java")

    canned = json.dumps(
        {
            "tasks": [
                {
                    "row_number": 2,
                    "name": "Hito MPP",
                    "wbs_code": "1",
                    "start_date": "2026-01-01",
                    "end_date": "2026-01-01",
                    "duration_days": 0,
                    "progress": 0,
                    "is_milestone": True,
                    "predecessors_raw": None,
                    "resources_raw": None,
                }
            ]
        }
    ).encode()

    def _fake_run(*args, **kwargs):
        # Post fix 2026-04-25: el wrapper Java escribe el JSON a un
        # archivo de salida (cmd[5]) en vez de stdout para evitar la
        # contaminación de logs de MPXJ/POI.
        cmd = args[0]
        if len(cmd) >= 6:
            with open(cmd[5], "wb") as f:
                f.write(canned)
        return subprocess.CompletedProcess(
            args=cmd, returncode=0, stdout=b"", stderr=b""
        )

    monkeypatch.setattr(mpp_parser.subprocess, "run", _fake_run)

    auth, proj_id = await _setup(client, db_session)
    files = {"file": ("plan.mpp", b"\x00MPP_BINARY", "application/vnd.ms-project")}
    r_prev = await client.post(
        f"/api/v1/projects/{proj_id}/tasks/import/preview",
        files=files,
        headers=auth["_authz"],
    )
    assert r_prev.status_code == 200, r_prev.text
    body = r_prev.json()
    assert body["source"] == "mpp"
    assert body["task_count"] == 1
    job_id = body["job_id"]

    r_conf = await client.post(
        f"/api/v1/projects/{proj_id}/tasks/import/{job_id}/confirm",
        json={"strategy": "replace"},
        headers=auth["_authz"],
    )
    assert r_conf.status_code == 200, r_conf.text
    assert r_conf.json()["imported"] == 1
    assert r_conf.json()["source"] == "mpp"


# --- TC-070.7 — Cross-user ownership: otro user del mismo tenant no puede confirmar ---


@pytest.mark.asyncio
async def test_tc070_7_cross_user_confirm_forbidden(client, db_session):
    from tests.factories import create_user as _mk_user

    auth, proj_id = await _setup(client, db_session)
    # Preview con user admin.
    data = _build_xlsx(SIMPLE_ROWS)
    files = {
        "file": (
            "plan.xlsx",
            data,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    }
    r_prev = await client.post(
        f"/api/v1/projects/{proj_id}/tasks/import/preview",
        files=files,
        headers=auth["_authz"],
    )
    job_id = r_prev.json()["job_id"]

    # Crear otro admin en el mismo tenant y loguearlo.
    from sqlalchemy import select

    from app.models.tenant import Tenant

    tenant = (await db_session.execute(select(Tenant).limit(1))).scalar_one()
    from app.models.role import Role
    role = (
        await db_session.execute(
            select(Role).where(Role.tenant_id == tenant.id).limit(1)
        )
    ).scalar_one()
    await _mk_user(
        db_session,
        tenant=tenant,
        username="admin2",
        email="admin2@acme.example.com",
        password="Str0ng-Admin-2!",
        roles=[role],
    )
    auth2 = await login(client, "admin2", "Str0ng-Admin-2!")

    r_conf = await client.post(
        f"/api/v1/projects/{proj_id}/tasks/import/{job_id}/confirm",
        json={"strategy": "replace"},
        headers=auth2["_authz"],
    )
    # Distinto user_id que el preview → 403 forbidden.
    assert r_conf.status_code == 403
