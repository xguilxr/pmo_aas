"""US-193 — plantilla profesional: detección automática de la fila de headers.

El layout pro trae encabezado del proyecto (filas 1-2) + KPIs (4-5) y
la tabla empieza en la fila 7. El parser busca la fila de headers (la
primera con alias de 'name') en las primeras 15 filas, ignora lo de
arriba y parsea los datos desde ahí — los archivos planos (header en
fila 1) siguen funcionando igual.
"""
from __future__ import annotations

import io
from datetime import date

from openpyxl import Workbook

from app.services.xlsx_task_parser import parse_xlsx


def _build_pro_layout() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Plan"
    # Encabezado del proyecto + KPIs (deben ignorarse).
    ws["A1"] = "T2M — Time2Market"
    ws["A2"] = "Sponsor"
    ws["B2"] = "L. Casados"
    ws["A4"] = "AVANCE GENERAL"
    ws["A5"] = "=ROUND(AVERAGEIF(C8:C60,1,G8:G60)*100,0)"
    # Fila 7: headers de la tabla.
    ws.append([])  # fila 6 vacía
    headers = [
        "WBS", "Tarea", "Nivel", "Inicio", "Fin", "Días", "Avance",
        "Estado", "Área", "Responsable", "Criticidad", "Hito",
        "Hito Relacionado", "Predecesoras", "Sucesoras",
    ]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=7, column=i, value=h)
    # Gantt a la derecha (headers de fecha — no deben mapear a nada).
    ws.cell(row=7, column=17, value=date(2026, 7, 6))
    # Datos desde la fila 8. Avance como fracción con formato %.
    data = [
        ("1", "Preparación", 1, date(2026, 7, 6), date(2026, 7, 17), 12,
         1.0, "Completada", "PMO", "D. Aguilar", "No", "No"),
        ("1.30", "Sub-fase treinta", 2, date(2026, 7, 8), date(2026, 7, 10),
         3, 0.45, "En progreso", "IT", "R. García", "Sí", "No"),
        ("1.30.1", "Detalle", 3, date(2026, 7, 9), date(2026, 7, 9), 1,
         0, "No iniciada", "IT", "", "No", "Sí"),
    ]
    for r_i, row in enumerate(data, start=8):
        for c_i, v in enumerate(row, start=1):
            cell = ws.cell(row=r_i, column=c_i, value=v)
            if c_i == 1:
                cell.number_format = "@"
            if c_i == 7:
                cell.number_format = "0%"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_pro_layout_header_detection_and_roundtrip():
    res = parse_xlsx(_build_pro_layout())
    assert res.header_row == 7
    # El encabezado/KPIs de arriba NO se convierten en tareas.
    assert [t.wbs_code for t in res.tasks] == ["1", "1.30", "1.30.1"]
    assert res.tasks[0].name == "Preparación"
    # % como fracción con formato % → 0-100; estados ES → enum.
    assert [t.progress for t in res.tasks] == [100, 45, 0]
    assert [t.status for t in res.tasks] == [
        "completed", "in_progress", "not_started",
    ]
    assert res.tasks[1].is_critical is True
    assert res.tasks[2].is_milestone is True
    # Números de fila reales (para warnings): datos desde la 8.
    assert res.tasks[0].row_number == 8


def test_flat_layout_still_row1():
    wb = Workbook()
    ws = wb.active
    ws.append(["WBS", "Nombre"])
    ws.append(["1", "Plana"])
    buf = io.BytesIO()
    wb.save(buf)
    res = parse_xlsx(buf.getvalue())
    assert res.header_row == 1
    assert [t.wbs_code for t in res.tasks] == ["1"]
