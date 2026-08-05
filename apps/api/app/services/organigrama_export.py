"""US-150 — Organigrama export (4 sheets) con openpyxl.

Genera un Excel con la estructura organizacional visible para el proyecto:
Áreas, Equipos, Roles y Recursos (actores). Mismo estilo que
`raid_export.py`: header bold con fondo del DS, freeze pane en fila 2 y
autosize best-effort. Los nombres (área/equipo/manager) se resuelven a
texto en el endpoint y se pasan ya formateados.
"""
from __future__ import annotations

from io import BytesIO
from typing import Any

from app.core.tipografia import FUENTE, aplicar_a_workbook

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _autosize(ws, ncols: int) -> None:
    from openpyxl.utils import get_column_letter

    for idx in range(1, ncols + 1):
        col_letter = get_column_letter(idx)
        max_len = 0
        for cell in ws[col_letter]:
            try:
                v = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(v))
            except Exception:
                continue
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 60)


def _write_sheet(wb, title: str, headers: list[str], rows: list[list[Any]]) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill

    ws = wb.create_sheet(title=title)
    header_font = Font(name=FUENTE, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_align = Alignment(vertical="center")

    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    for row in rows:
        ws.append(["" if v is None else v for v in row])

    ws.freeze_panes = "A2"
    _autosize(ws, len(headers))


# --- US-186: hojas de utilización (FTE + uso mensual con alertas) -----------

FTE_HEADERS = [
    "Recurso", "Función", "Puesto", "Área", "Equipo", "Manager", "Tipo",
    "Clave", "Capacidad %", "% FTE en scope (mes actual)",
    "% FTE total tenant (mes actual)", "Proyectos en scope",
]

# Alertas por diseño (owner 2026-07-09): ≥80% amarillo, >100% rojo.
UTIL_YELLOW_FROM = 80
UTIL_RED_FROM = 100


def _util_row(r: dict[str, Any]) -> list[Any]:
    return [
        r["name"], r.get("discipline") or "", r.get("job_title") or "",
        r.get("area") or "", r.get("team") or "", r.get("manager") or "",
        r.get("resource_type") or "", "Sí" if r.get("is_key_resource") else "",
        r.get("capacity_pct"), r.get("scope_current_pct"),
        r.get("tenant_current_pct"), r.get("projects_count"),
    ]


def _write_monthly_sheet(
    wb, title: str, months: list[str], rows: list[dict[str, Any]]
) -> None:
    """Hoja "Uso mensual": Recurso × Mes. Fill amarillo si el mes ≥80%,
    rojo si >100% + columna 'Meses en alerta'."""
    from openpyxl.styles import Font, PatternFill

    headers = ["Recurso", *months, "Meses en alerta"]
    data_rows = [[r["name"], *r["per_month"], r["alert_months"]] for r in rows]
    _write_sheet(wb, title, headers, data_rows)

    ws = wb[title]
    yellow = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    red = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    red_font = Font(name=FUENTE, color="B91C1C", bold=True)
    yellow_font = Font(name=FUENTE, color="92400E")
    for i, r in enumerate(rows, start=2):  # fila 1 = header
        for j, val in enumerate(r["per_month"], start=2):  # col 1 = nombre
            cell = ws.cell(row=i, column=j)
            if val > UTIL_RED_FROM:
                cell.fill = red
                cell.font = red_font
            elif val >= UTIL_YELLOW_FROM:
                cell.fill = yellow
                cell.font = yellow_font


def export_utilizacion_xlsx(
    *, months: list[str], rows: list[dict[str, Any]]
) -> bytes:
    """US-186 — organigrama de utilización para programa/organización/
    tenant: hoja "Organigrama" (recursos activos + FTE) + "Uso mensual"."""
    from openpyxl import Workbook

    wb = Workbook()

    # ENH-202: Helvetica en todas las celdas sin estilo propio.

    aplicar_a_workbook(wb)
    default_ws = wb.active
    if default_ws is not None:
        wb.remove(default_ws)
    _write_sheet(wb, "Organigrama", FTE_HEADERS, [_util_row(r) for r in rows])
    _write_monthly_sheet(wb, "Uso mensual", months, rows)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_organigrama_xlsx(
    *,
    areas_rows: list[list[Any]],
    teams_rows: list[list[Any]],
    roles_rows: list[list[Any]],
    recursos_rows: list[list[Any]],
    utilization_months: list[str] | None = None,
    utilization_rows: list[dict[str, Any]] | None = None,
) -> bytes:
    """Devuelve bytes XLSX con 4 sheets (Áreas/Equipos/Roles/Recursos).

    Cada `*_rows` es una lista de filas ya formateadas (nombres resueltos),
    en el orden de las columnas declaradas abajo. Sheets vacíos se incluyen
    igual con sólo el header.
    """
    from openpyxl import Workbook

    wb = Workbook()

    # ENH-202: Helvetica en todas las celdas sin estilo propio.

    aplicar_a_workbook(wb)
    default_ws = wb.active
    if default_ws is not None:
        wb.remove(default_ws)

    _write_sheet(
        wb, "Áreas",
        ["ID", "Nombre", "Descripción", "Líder", "Activa"],
        areas_rows,
    )
    _write_sheet(
        wb, "Equipos",
        ["ID", "Área", "Nombre", "Descripción", "Activo"],
        teams_rows,
    )
    _write_sheet(
        wb, "Roles",
        ["ID", "Nombre", "Descripción", "Activo"],
        roles_rows,
    )
    _write_sheet(
        wb, "Recursos",
        [
            "ID", "Nombre", "Equipo", "Área", "Puesto", "Empresa",
            "Email", "Teléfono", "Manager", "Activo",
        ],
        recursos_rows,
    )

    # US-186: hojas de utilización (FTE + uso mensual) si el caller las
    # calculó (scope proyecto: participaciones activas con FTE%).
    if utilization_months is not None and utilization_rows is not None:
        _write_sheet(
            wb, "Recursos (FTE)", FTE_HEADERS,
            [_util_row(r) for r in utilization_rows],
        )
        _write_monthly_sheet(wb, "Uso mensual", utilization_months, utilization_rows)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
