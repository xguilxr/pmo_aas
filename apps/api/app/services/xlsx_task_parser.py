"""US-067 — Parser XLSX de tareas estilo MS Project / Excel genérico.

Auto-detecta columnas con headers en español/inglés comunes:
- Nombre: "Nombre", "Tarea", "Task Name", "Name"
- WBS: "WBS", "EDT", "Código"
- Inicio: "Inicio", "Start", "Fecha inicio", "Start Date"
- Fin: "Fin", "Finish", "Fecha fin", "Finish Date", "End Date"
- Duración: "Duración", "Duration", "Días", "Duración (días)"
- Avance: "% completado", "Progress", "%Complete", "Avance", "Avance (%)"
- Hito: "Hito", "Milestone", "Es hito"
- Predecesores: "Predecesoras", "Predecessors"
- Recursos: "Recursos", "Resources", "Responsable", "Resource Names"

Devuelve una lista de `ParsedTask` lista para persistir. Errores por fila
se acumulan en `errors` con el número de fila y el detalle.
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import date, datetime
from io import BytesIO

HEADER_ALIASES: dict[str, list[str]] = {
    "name": ["nombre", "tarea", "task name", "name", "nombre de tarea"],
    "wbs": ["wbs", "edt", "codigo", "código"],
    "start_date": ["inicio", "start", "fecha inicio", "start date", "fecha de inicio"],
    "end_date": ["fin", "finish", "fecha fin", "finish date", "end date", "fecha de fin"],
    "duration_days": [
        "duracion",
        "duración",
        "duration",
        "días",
        "dias",
        "duración (días)",
        "duracion (dias)",
    ],
    "progress": [
        "% completado",
        "%completado",
        "progress",
        "%complete",
        "avance",
        "avance (%)",
        "% avance",
    ],
    "is_milestone": ["hito", "milestone", "es hito"],
    # ENH-134: la columna "Criticidad" de la plantilla V1 es booleana
    # (Sí/No) → mapea a is_critical. El enum legacy queda accesible solo
    # vía mapeo manual ("prioridad criticidad").
    "criticality": ["prioridad criticidad"],
    "is_critical": [
        "criticidad",
        "criticality",
        "is_critical",
        "es critico",
        "es crítico",
        "critico",
        "crítico",
    ],
    "related_milestone": [
        "hito relacionado",
        "related milestone",
        "milestone relacionado",
    ],
    "predecessors": ["predecesoras", "predecessors", "predecessoras"],
    # ENH-134: área responsable (se resuelve a area_id en el confirm).
    "area": [
        "área responsable",
        "area responsable",
        "área",
        "area",
        "área responsable ",
    ],
    "resources": [
        "recursos",
        "resources",
        "responsable",
        "resource names",
        "responsables",
    ],
}


@dataclass
class ParsedTask:
    row_number: int
    name: str
    wbs: str | None = None
    start_date: date | None = None
    end_date: date | None = None
    duration_days: int | None = None
    progress: int = 0
    is_milestone: bool = False
    # US-096: criticidad + hito relacionado opcionales en plantilla.
    criticality: str | None = None
    # ENH-097: boolean explicito. None = no presente en plantilla (caller
    # debe derivarlo de criticality).
    is_critical: bool | None = None
    related_milestone_wbs: str | None = None
    predecessors_raw: str | None = None
    resources_raw: str | None = None
    # ENH-134: nombre del área responsable (texto). Se resuelve a area_id
    # contra las áreas del proyecto en el confirm.
    area_raw: str | None = None


@dataclass
class XlsxParseResult:
    tasks: list[ParsedTask] = field(default_factory=list)
    errors: list[dict] = field(default_factory=list)
    # BUG-088: avisos no bloqueantes (ej. WBS numérico sin formato de
    # texto). Shape: {code, message, count?, rows?}. El wizard los
    # muestra en el preview; el import no se detiene por ellos.
    warnings: list[dict] = field(default_factory=list)
    columns_detected: dict[str, int] = field(default_factory=dict)
    # US-070: hojas disponibles en el workbook. Vacío para CSV/MPP;
    # solo el parser XLSX lo puebla. El wizard lo usa en el step 2
    # (sheet selector) cuando hay más de una hoja.
    sheets: list[str] = field(default_factory=list)
    # US-070: primeras N filas crudas (incluye header) para que el
    # wizard pueda renderizar el preview + re-mapeo de columnas antes
    # del confirm. Cada elemento es una lista de celdas tal como vino
    # del parser (sin coerción).
    sample_rows: list[list[object]] = field(default_factory=list)
    # US-070: nombre de la hoja efectivamente parseada. En Excel con
    # varias hojas el caller puede elegir; en CSV/MPP queda None.
    sheet_used: str | None = None


def _norm(s: object) -> str:
    if s is None:
        return ""
    return str(s).strip().lower()


def _text(v: object) -> str | None:
    """Strip preservando mayúsculas (a diferencia de `_norm`). Para
    campos donde el case importa: recursos/responsables, predecesoras."""
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def _decimal_places(fmt: str | None) -> int | None:
    """Decimales forzados por un `number_format` de Excel ('0.00' → 2).

    None cuando el formato no fija decimales (General, texto '@', o
    formato desconocido); 0 cuando es entero explícito ('0', '#,##0')."""
    if not fmt:
        return None
    primary = re.sub(r"\[[^\]]*\]", "", fmt.split(";")[0]).strip()
    if not primary or primary.lower() == "general" or primary == "@":
        return None
    if "." not in primary:
        return 0
    forced = 0
    for ch in primary.split(".", 1)[1]:
        if ch == "0":
            forced += 1
        elif ch in "#?":
            continue
        else:
            break
    return forced


def _wbs_text(value: object, number_format: str | None = None) -> str | None:
    """BUG-088 — texto fiel de un código WBS/EDT.

    Excel guarda '1.30' tipeado en celda numérica como el float 1.3;
    el str() directo colapsaba 1.30 → '1.3' (colisión con el 1.3 real
    y sub-tareas 1.30.x huérfanas). Reglas:
    - string → strip (sin lowercasing: '1.A' se preserva).
    - float con formato decimal ('0.00') → respeta esos decimales
      (1.3 + '0.00' → '1.30').
    - float sin formato → representación mínima (1.3 → '1.3'); entero
      exacto → sin '.0' (2.0 → '2').
    """
    if value is None:
        return None
    if isinstance(value, float) and not isinstance(value, bool):
        decs = _decimal_places(number_format)
        if decs is not None and decs > 0:
            return f"{value:.{decs}f}"
        if value.is_integer():
            return str(int(value))
        return repr(value)
    s = str(value).strip()
    return s or None


def _detect_headers(header_row: list[object]) -> dict[str, int]:
    """Mapea cada header canónico → índice de columna. Columnas no
    reconocidas se ignoran."""
    detected: dict[str, int] = {}
    for idx, cell in enumerate(header_row):
        label = _norm(cell)
        if not label:
            continue
        for canonical, aliases in HEADER_ALIASES.items():
            if label in aliases:
                detected.setdefault(canonical, idx)
                break
    return detected


def _coerce_date(v: object) -> date | None:
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    try:
        return datetime.fromisoformat(str(v)).date()
    except ValueError:
        return None


def _coerce_int(v: object, default: int = 0) -> int:
    if v is None or v == "":
        return default
    try:
        return int(float(str(v)))
    except ValueError:
        return default


def _coerce_progress(v: object, *, is_percent_format: bool = False) -> int:
    """Acepta 0-100, 0-1 (fracción) o texto '45%'.

    BUG-081: cuando la celda de Excel tiene `number_format` de porcentaje,
    openpyxl entrega la *fracción* (1.0==100%, 0.3==30%, 0.5==50%). El
    heurístico viejo escalaba ×100 solo si el string tenía ".", así que
    100% (que openpyxl devuelve como el entero ``1``, sin decimales) se
    quedaba en 1% — exactamente el síntoma reportado. Con
    ``is_percent_format=True`` siempre escalamos ×100.
    """
    if v is None or v == "":
        return 0
    s = str(v).strip()
    had_pct_sign = "%" in s
    s = s.replace("%", "").strip()
    try:
        n = float(s)
    except ValueError:
        return 0
    if is_percent_format:
        # Celda %-formateada → openpyxl da la fracción 0..1 (incl. 1==100%).
        scaled = n * 100
        if scaled > 100 and 0 <= n <= 100:
            # BUG-089: celda con formato % pero valor entero tipeado
            # (45 → Excel muestra 4500%). Escalar daba 4500 → clamp 100
            # y TODO el plan quedaba en 100%. El usuario quiso 45%: se
            # usa el valor literal.
            pass
        else:
            n = scaled
    elif 0 < n < 1 and not had_pct_sign:
        # Fracción literal sin formato (ej. CSV "0.45" == 45%).
        n = n * 100
    return max(0, min(100, round(n)))


def _scan_column_formats(data: bytes, sheet_used: str, col_idx: int) -> dict[int, str]:
    """BUG-088: `{row_number: number_format}` de la columna `col_idx`
    (0-based). El workbook principal se abre con `values_only`, que no
    expone formatos; este segundo pase read-only los recupera (mismo
    patrón que `_column_is_percent_format` de BUG-081). Devuelve {}
    ante cualquier problema (degradación segura)."""
    from openpyxl import load_workbook

    out: dict[int, str] = {}
    try:
        wb = load_workbook(BytesIO(data), read_only=True)
    except Exception:
        return out
    try:
        ws = wb[sheet_used] if sheet_used in wb.sheetnames else wb.active
        if ws is None:
            return out
        for row in ws.iter_rows(min_row=2, min_col=col_idx + 1, max_col=col_idx + 1):
            if row and row[0].number_format:
                out[row[0].row] = row[0].number_format
        return out
    except Exception:
        return out
    finally:
        try:
            wb.close()
        except Exception:
            pass


def _coerce_bool(v: object) -> bool:
    s = _norm(v)
    return s in {"sí", "si", "yes", "true", "1", "x", "✓"}


SAMPLE_ROW_LIMIT = 10


def parse_xlsx(
    data: bytes,
    sheet: str | None = None,
    columns_override: dict[str, int] | None = None,
    strict: bool = True,
) -> XlsxParseResult:
    """Parsea XLSX en-memoria y devuelve ParsedTask + errors.

    US-070: acepta `sheet` opcional para elegir hoja específica,
    `columns_override` para mapeo manual (reemplaza totalmente la
    auto-detección) y `strict` que controla qué pasa cuando no hay
    columna `name`:

    - `strict=True` (default): raise ValueError. Caller se rompe si
      el archivo no tiene el header esperado.
    - `strict=False`: devuelve `XlsxParseResult` con `tasks=[]` y
      `sample_rows[]` poblado. Útil para el preview del wizard donde
      el usuario va a re-mappear columnas manualmente.
    """
    from openpyxl import load_workbook

    result = XlsxParseResult()
    try:
        wb = load_workbook(BytesIO(data), data_only=True, read_only=True)
    except Exception as exc:
        raise ValueError(f"archivo XLSX inválido: {exc}") from exc

    result.sheets = list(wb.sheetnames)
    if sheet is not None:
        if sheet not in result.sheets:
            raise ValueError(
                f"hoja '{sheet}' no existe en el workbook (disponibles: "
                f"{', '.join(result.sheets)})"
            )
        ws = wb[sheet]
    else:
        ws = wb.active
    if ws is None:
        raise ValueError("workbook sin hojas")
    result.sheet_used = ws.title

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = list(next(rows_iter))
    except StopIteration:
        return result

    # Sample para el wizard: header + hasta N data rows. Se guarda
    # antes de iterar el resto para no consumir el iterator.
    result.sample_rows.append(list(header_row))

    columns = (
        dict(columns_override)
        if columns_override is not None
        else _detect_headers(header_row)
    )
    result.columns_detected = columns
    if "name" not in columns:
        if strict:
            raise ValueError(
                "Falta mapear la columna obligatoria 'Nombre'. Asegurá que la "
                "hoja tenga headers estándar en la fila 1 o enviá un mapping "
                "manual."
            )
        # Modo preview: seguimos poblando sample_rows y devolvemos el
        # result parcial para que el wizard pueda mostrarle los headers
        # al usuario y que mapee manualmente.
        for row in rows_iter:
            if row is None:
                continue
            if len(result.sample_rows) <= SAMPLE_ROW_LIMIT:
                result.sample_rows.append(list(row))
            else:
                break
        return result

    # BUG-081 + BUG-089: formato % de la columna de avance POR CELDA
    # (openpyxl entrega fracciones para celdas %-formateadas). La
    # detección por columna escalaba también celdas planas de columnas
    # con formatos mixtos → todo terminaba en 100%.
    progress_formats: dict[int, str] = (
        _scan_column_formats(data, result.sheet_used, columns["progress"])
        if "progress" in columns
        else {}
    )
    progress_literal_rows: list[int] = []

    # BUG-088: formatos de la columna WBS para reconstruir el texto fiel
    # de celdas numéricas (1.30 con formato '0.00' ≠ float 1.3).
    wbs_formats: dict[int, str] = (
        _scan_column_formats(data, result.sheet_used, columns["wbs"])
        if "wbs" in columns
        else {}
    )
    wbs_general_rows: list[int] = []

    for offset, row in enumerate(rows_iter, start=2):
        if row is None:
            continue
        # US-070: acumulá sample rows hasta el límite — antes del
        # filtro por `name` para que el wizard pueda mostrar incluso
        # filas que el parser actual descarta (ej. resumenes en blanco).
        if len(result.sample_rows) <= SAMPLE_ROW_LIMIT:
            result.sample_rows.append(list(row))
        name_cell = row[columns["name"]] if columns["name"] < len(row) else None
        name = _norm(name_cell)
        if not name:
            continue
        # BUG-088: celda WBS numérica con fracción y sin formato decimal
        # explícito → los ceros finales son irrecuperables. Warning.
        raw_wbs = (
            row[columns["wbs"]]
            if "wbs" in columns and columns["wbs"] < len(row)
            else None
        )
        if (
            isinstance(raw_wbs, float)
            and not raw_wbs.is_integer()
            and not _decimal_places(wbs_formats.get(offset))
        ):
            wbs_general_rows.append(offset)
        # BUG-089: celda de avance %-formateada con entero tipeado
        # (45 en vez de 45%) → se interpreta como valor literal 0-100.
        raw_prog = (
            row[columns["progress"]]
            if "progress" in columns and columns["progress"] < len(row)
            else None
        )
        prog_is_pct = "%" in (progress_formats.get(offset) or "")
        if (
            prog_is_pct
            and isinstance(raw_prog, (int, float))
            and not isinstance(raw_prog, bool)
            and 1 < float(raw_prog) <= 100
        ):
            progress_literal_rows.append(offset)
        try:
            task = ParsedTask(
                row_number=offset,
                name=str(name_cell).strip(),
                wbs=_wbs_text(raw_wbs, wbs_formats.get(offset)),
                start_date=_coerce_date(row[columns["start_date"]])
                if "start_date" in columns and columns["start_date"] < len(row)
                else None,
                end_date=_coerce_date(row[columns["end_date"]])
                if "end_date" in columns and columns["end_date"] < len(row)
                else None,
                duration_days=_coerce_int(row[columns["duration_days"]], default=0)
                if "duration_days" in columns and columns["duration_days"] < len(row)
                else None,
                progress=_coerce_progress(raw_prog, is_percent_format=prog_is_pct)
                if "progress" in columns and columns["progress"] < len(row)
                else 0,
                is_milestone=_coerce_bool(row[columns["is_milestone"]])
                if "is_milestone" in columns and columns["is_milestone"] < len(row)
                else False,
                criticality=(_norm(row[columns["criticality"]]) or None)
                if "criticality" in columns and columns["criticality"] < len(row)
                else None,
                is_critical=_coerce_bool(row[columns["is_critical"]])
                if "is_critical" in columns and columns["is_critical"] < len(row)
                else None,
                related_milestone_wbs=_wbs_text(row[columns["related_milestone"]])
                if "related_milestone" in columns
                and columns["related_milestone"] < len(row)
                else None,
                predecessors_raw=_wbs_text(row[columns["predecessors"]])
                if "predecessors" in columns and columns["predecessors"] < len(row)
                else None,
                resources_raw=_text(row[columns["resources"]])
                if "resources" in columns and columns["resources"] < len(row)
                else None,
                area_raw=(str(row[columns["area"]]).strip() or None)
                if "area" in columns
                and columns["area"] < len(row)
                and row[columns["area"]] is not None
                else None,
            )
            result.tasks.append(task)
        except Exception as exc:
            result.errors.append({"row": offset, "error": str(exc)})

    if progress_literal_rows:
        result.warnings.append(
            {
                "code": "PROGRESS_PCT_AS_INTEGER",
                "count": len(progress_literal_rows),
                "rows": progress_literal_rows[:20],
                "message": (
                    "La columna Avance tiene formato de porcentaje pero "
                    "valores enteros (ej. 45 en vez de 45%): se "
                    "interpretaron como 0-100. Verificá los avances en la "
                    "vista previa."
                ),
            }
        )
    if wbs_general_rows:
        result.warnings.append(
            {
                "code": "WBS_NUMERIC_GENERAL",
                "count": len(wbs_general_rows),
                "rows": wbs_general_rows[:20],
                "message": (
                    "La columna WBS tiene celdas numéricas sin formato de "
                    "texto: Excel pierde los ceros finales (1.30 se lee "
                    "como 1.3). Formateá la columna WBS como Texto para "
                    "preservar la numeración."
                ),
            }
        )

    return result
