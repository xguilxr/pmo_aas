"""Change Requests export — ENH-186.

Genera un Excel de **1 hoja "Cambios"** con columnas en español. Mismo
estilo que `raid_export.py` (ENH-152): header bold con fondo del DS,
freeze pane en fila 2 y autosize best-effort. Las filas llegan ya
formateadas (nombres de solicitante/aprobador resueltos a texto en el
endpoint), siguiendo el patrón de `organigrama_export.py`.
"""
from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
from typing import Any

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

CHANGE_HEADERS: list[str] = [
    "Folio",
    "Título",
    "Tipo",
    "Estado",
    "Solicitado por",
    "Fecha solicitud",
    "Aprobado por",
    "Fecha aprobación",
    "Impacto",
]

# Consistente con CHANGE_TYPE_LABEL / CHANGE_STATUS_LABEL en
# apps/web/lib/api/modules.ts. Incluye estados legacy del flujo EP019
# (draft/pending_approval) por si quedan filas en esos estados.
CHANGE_TYPE_ES: dict[str, str] = {
    "scope": "Alcance",
    "time": "Tiempo",
    "cost": "Costo",
    "resource": "Recursos",
}

CHANGE_STATUS_ES: dict[str, str] = {
    "draft": "Borrador",
    "in_review": "En revisión",
    "pending_approval": "Pendiente de aprobación",
    "approved": "Aprobado",
    "rejected": "Rechazado",
    "implemented": "Implementado",
    "cancelled": "Cancelado",
}


def build_change_rows(
    changes: list[Any],
    user_names: dict[str, str],
) -> list[list[Any]]:
    """Filas de la hoja Cambios. `user_names` resuelve requested_by /
    approved_by (UUID de User) a nombre para mostrar texto en vez de UUID."""
    rows: list[list[Any]] = []
    for c in changes:
        requested_at = c.requested_at
        requested_date = requested_at.date() if isinstance(requested_at, datetime) else requested_at
        approved_at = c.approved_at
        approved_date = approved_at.date() if isinstance(approved_at, datetime) else (approved_at or "")
        rows.append(
            [
                c.folio,
                c.title,
                CHANGE_TYPE_ES.get(c.type, c.type),
                CHANGE_STATUS_ES.get(c.status, c.status),
                user_names.get(str(c.requested_by), "") if c.requested_by else "",
                requested_date or "",
                user_names.get(str(c.approved_by), "") if c.approved_by else "",
                approved_date,
                c.impact or "",
            ]
        )
    return rows


def _cell_value(value: Any) -> Any:
    """Convierte tipos no-Excel-friendly. tz-aware datetimes → naive (Excel
    no acepta tz). Listas/dicts → string."""
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.replace(tzinfo=None) if value.tzinfo else value
    if isinstance(value, date):
        return value
    if isinstance(value, (list, dict)):
        return ", ".join(str(x) for x in value) if isinstance(value, list) else str(value)
    return value


def _autosize(ws, n_cols: int) -> None:
    """Best-effort: ancho = max longitud de la columna, cap a 60."""
    from openpyxl.utils import get_column_letter

    for idx in range(1, n_cols + 1):
        col_letter = get_column_letter(idx)
        max_len = 0
        for cell in ws[col_letter]:
            try:
                v = "" if cell.value is None else str(cell.value)
                if len(v) > max_len:
                    max_len = len(v)
            except Exception:
                continue
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)


def _write_sheet(wb, title: str, headers: list[str], rows: list[list[Any]]) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill

    ws = wb.create_sheet(title=title)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_align = Alignment(vertical="center")

    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    for row in rows:
        ws.append([_cell_value(v) for v in row])

    ws.freeze_panes = "A2"
    _autosize(ws, len(headers))


def export_changes_xlsx(rows: list[list[Any]]) -> bytes:
    """Devuelve bytes XLSX con 1 hoja "Cambios". Si no hay filas, incluye
    igual el header."""
    from openpyxl import Workbook

    wb = Workbook()
    default_ws = wb.active
    if default_ws is not None:
        wb.remove(default_ws)
    _write_sheet(wb, "Cambios", CHANGE_HEADERS, rows)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
