"""RAID export — 4 hojas RAID en español (ENH-152).

Genera un Excel con 4 hojas dedicadas: **Riesgos / Acciones / Incidencias /
Decisiones**. Header row en negrita con fondo del DS, freeze pane en fila 2 y
autosize best-effort. Hojas vacías se incluyen igual con header.

Las filas llegan ya formateadas (nombres de área y responsable resueltos a
texto en el endpoint, vía `build_risk_rows` / `build_issue_rows`), siguiendo el
patrón de `organigrama_export`. Es el **mismo archivo** para el botón de `/raid`
y el del módulo Documentos.

Reusa los modelos `Risk` e `Issue` de `app.models.modules`. Sin cambio de schema:
los 4 tipos RAID son `Risk` + `Issue.type` (action / issue / decision).
"""
from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
from typing import Any

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

# Columnas por hoja (ENH-152). Riesgos usa "Severidad"; los AID, "Prioridad".
RISK_HEADERS: list[str] = [
    "Folio", "Título", "Descripción", "Severidad", "Estado",
    "Responsable área", "Responsable", "Fecha creación",
]
AID_HEADERS: list[str] = [
    "Folio", "Título", "Descripción", "Prioridad", "Estado",
    "Responsable área", "Responsable", "Fecha creación",
]

# Labels ES de estado para que el export sea presentable.
RISK_STATUS_ES: dict[str, str] = {
    "identified": "Identificado",
    "analyzing": "En análisis",
    "mitigating": "Mitigando",
    "materialized": "Materializado",
    "closed": "Cerrado",
}
ISSUE_STATUS_ES: dict[str, str] = {
    "open": "Abierta",
    "in_progress": "En progreso",
    "resolved": "Resuelta",
    "closed": "Cerrada",
}


def _responsible(
    owner_actor_id: Any,
    owner_id: Any,
    actor_names: dict[str, str],
    user_names: dict[str, str],
) -> str:
    """Responsable = nombre del Actor (catálogo) con fallback al Usuario."""
    if owner_actor_id and str(owner_actor_id) in actor_names:
        return actor_names[str(owner_actor_id)]
    if owner_id and str(owner_id) in user_names:
        return user_names[str(owner_id)]
    return ""


def _area(area_id: Any, area_names: dict[str, str]) -> str:
    return area_names.get(str(area_id), "") if area_id else ""


def build_risk_rows(
    risks: list[Any],
    area_names: dict[str, str],
    actor_names: dict[str, str],
    user_names: dict[str, str],
) -> list[list[Any]]:
    """Filas de la hoja Riesgos (Severidad + fecha de identificación)."""
    return [
        [
            r.folio,
            r.title,
            r.description or "",
            r.severity if r.severity is not None else "",
            RISK_STATUS_ES.get(r.status, r.status),
            _area(r.area_id, area_names),
            _responsible(r.owner_actor_id, r.owner_id, actor_names, user_names),
            r.identified_at if r.identified_at else "",
        ]
        for r in risks
    ]


def build_issue_rows(
    issues: list[Any],
    area_names: dict[str, str],
    actor_names: dict[str, str],
    user_names: dict[str, str],
) -> list[list[Any]]:
    """Filas de las hojas Acciones / Incidencias / Decisiones (Prioridad +
    fecha de reporte). El tipo se filtra antes de llamar (action/issue/decision)."""
    return [
        [
            i.folio,
            i.title,
            i.description or "",
            i.priority if i.priority is not None else "",
            ISSUE_STATUS_ES.get(i.status, i.status),
            _area(i.area_id, area_names),
            _responsible(i.owner_actor_id, i.owner_id, actor_names, user_names),
            i.reported_at.date() if i.reported_at else "",
        ]
        for i in issues
    ]


def _cell_value(value: Any) -> Any:
    """Convierte tipos no-Excel-friendly. tz-aware datetimes → naive (Excel
    no acepta tz). Listas/dicts → string."""
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.replace(tzinfo=None) if value.tzinfo else value
    if isinstance(value, date):
        return value
    if isinstance(value, (list, dict)):
        return ", ".join(str(x) for x in value) if isinstance(value, list) else str(value)
    return value


def _autosize(ws, n_cols: int) -> None:
    """Best-effort: ancho = max longitud de la columna, cap a 60."""
    from openpyxl.utils import get_column_letter

    for idx in range(1, n_cols + 1):
        col_letter = get_column_letter(idx)
        max_len = 0
        for cell in ws[col_letter]:
            try:
                v = "" if cell.value is None else str(cell.value)
                if len(v) > max_len:
                    max_len = len(v)
            except Exception:
                continue
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)


def _write_sheet(wb, title: str, headers: list[str], rows: list[list[Any]]) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill

    ws = wb.create_sheet(title=title)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_align = Alignment(vertical="center")

    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    for row in rows:
        ws.append([_cell_value(v) for v in row])

    # Freeze pane: fila 2 (header sticky).
    ws.freeze_panes = "A2"
    _autosize(ws, len(headers))


def export_single_sheet_xlsx(
    *, title: str, headers: list[str], rows: list[list[Any]]
) -> bytes:
    """ENH-168: XLSX de UNA sola hoja para export individual por tipo
    (Riesgos / Acciones / Incidencias / Decisiones)."""
    from openpyxl import Workbook

    wb = Workbook()
    default_ws = wb.active
    if default_ws is not None:
        wb.remove(default_ws)
    _write_sheet(wb, title, headers, rows)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_raid_xlsx(
    *,
    risks_rows: list[list[Any]],
    actions_rows: list[list[Any]],
    incidents_rows: list[list[Any]],
    decisions_rows: list[list[Any]],
) -> bytes:
    """Devuelve bytes XLSX con 4 hojas RAID en español. Hojas vacías se
    incluyen igual con sólo el header row."""
    from openpyxl import Workbook

    wb = Workbook()
    # Workbook trae 1 sheet default; lo eliminamos para controlar el orden.
    default_ws = wb.active
    if default_ws is not None:
        wb.remove(default_ws)

    _write_sheet(wb, "Riesgos", RISK_HEADERS, risks_rows)
    _write_sheet(wb, "Acciones", AID_HEADERS, actions_rows)
    _write_sheet(wb, "Incidencias", AID_HEADERS, incidents_rows)
    _write_sheet(wb, "Decisiones", AID_HEADERS, decisions_rows)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
