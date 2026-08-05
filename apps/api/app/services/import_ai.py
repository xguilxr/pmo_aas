"""US-188 — Import inteligente de planes con IA.

Tres niveles sobre la infra de ENH-053 (LLM del tenant vía
`generate_for_tenant`, gated por `tenant.ai_mode`; sin IA todo degrada
a la heurística y el import sigue funcionando):

1. **Mapeo por contenido** — `suggest_column_mapping` acepta filas de
   muestra para decidir por los VALORES, no solo por el header (vive en
   `import_mapping_suggest.py`).
2. **Normalización de valores** — estados libres ("casi lista",
   "80% done") → enum canónico; responsables que el fuzzy-match no
   resolvió → actores del pool. Se aplica en el confirm.
3. **Estructura** — un archivo "sucio" (headers crípticos, secciones,
   indentación en vez de WBS) se convierte en una propuesta de plan
   completa vía structured output; el usuario la revisa en el preview
   del wizard antes de confirmar.

Todas las funciones son best-effort: ante cualquier error del LLM
devuelven vacío y el caller sigue con la heurística.
"""
from __future__ import annotations

import csv
import io
import json
import logging

from app.services.ai.prompt_builder import build_system_prompt
from app.services.ai.provider import generate_for_tenant
from app.services.ai.tenant_ai import TenantAIConfig
from app.services.ai.untrusted import envolver_no_confiable
from app.services.xlsx_task_parser import (
    ParsedTask,
    _coerce_date,
    _coerce_status,
)

log = logging.getLogger(__name__)

STATUS_ENUM = ("not_started", "in_progress", "completed", "on_hold", "cancelled")

# Límites para no mandar archivos completos al LLM.
MAX_STRUCTURE_ROWS = 200
MAX_STRUCTURE_COLS = 12
MAX_VALUES_PER_CALL = 40


def extract_raw_rows(
    source: str,
    data: bytes,
    sheet: str | None = None,
    limit: int = MAX_STRUCTURE_ROWS,
) -> list[list[str | None]]:
    """Filas crudas (como strings) de un XLSX/CSV para el nivel 3.
    Incluye TODO desde la fila 1 — la gracia es que la IA decida qué es
    header, sección o dato."""
    rows: list[list[str | None]] = []
    if source == "xlsx":
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
        try:
            ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active
            if ws is None:
                return rows
            for row in ws.iter_rows(values_only=True):
                rows.append([None if c is None else str(c) for c in row])
                if len(rows) >= limit:
                    break
        finally:
            wb.close()
        return rows
    if source == "csv":
        from app.services.csv_task_parser import _decode, _sniff_dialect

        text = _decode(data)
        reader = csv.reader(io.StringIO(text), dialect=_sniff_dialect(text[:4096]))
        for row in reader:
            rows.append([c if c != "" else None for c in row])
            if len(rows) >= limit:
                break
        return rows
    return rows


def _extract_json(text: str) -> dict | list | None:
    """Extrae JSON aunque el LLM lo envuelva en backticks o prosa."""
    s = (text or "").strip()
    if s.startswith("```"):
        s = "\n".join(line for line in s.splitlines() if not line.startswith("```"))
    for open_ch, close_ch in (("{", "}"), ("[", "]")):
        start = s.find(open_ch)
        end = s.rfind(close_ch)
        if start != -1 and end > start:
            try:
                return json.loads(s[start : end + 1])
            except json.JSONDecodeError:
                continue
    return None


async def ai_normalize_statuses(
    values: list[str],
    *,
    tenant_cfg: TenantAIConfig,
    tenant_id: str | None = None,
) -> dict[str, str]:
    """Nivel 2 — mapea estados libres al enum canónico. Devuelve solo
    los que el LLM resolvió con un valor válido del enum."""
    values = [v for v in dict.fromkeys(values) if v][:MAX_VALUES_PER_CALL]
    if not values or tenant_cfg.mode == "disabled":
        return {}
    system = (
        "You map free-text task status values (Spanish or English) to one "
        f"of this enum: {', '.join(STATUS_ENUM)}. Reply ONLY strict JSON "
        '{"<raw value>": "<enum or null>"}. Use null when genuinely '
        "ambiguous. No prose."
    )
    try:
        res = await generate_for_tenant(
            envolver_no_confiable(
                json.dumps({"values": values}, ensure_ascii=False),
                origen="valores de estado del archivo subido por el usuario",
            ),
            system=build_system_prompt(system, None),
            tenant_ai_mode=tenant_cfg.mode,
            byo_config=tenant_cfg.byo,
            tenant_id=tenant_id,
            json_mode=True,
        )
        parsed = _extract_json(res.text)
    except Exception as exc:
        log.info("ai_normalize_statuses fallo: %s", exc)
        return {}
    if not isinstance(parsed, dict):
        return {}
    return {
        raw: mapped
        for raw, mapped in parsed.items()
        if raw in values and isinstance(mapped, str) and mapped in STATUS_ENUM
    }


async def ai_match_resources(
    values: list[str],
    actor_names: list[str],
    *,
    tenant_cfg: TenantAIConfig,
    tenant_id: str | None = None,
) -> dict[str, str]:
    """Nivel 2 — matchea responsables libres ("J. Pérez", "juan p")
    contra los nombres del pool de recursos. Devuelve {raw: nombre del
    pool} solo para matches que el LLM da por seguros."""
    values = [v for v in dict.fromkeys(values) if v][:MAX_VALUES_PER_CALL]
    actor_names = [a for a in actor_names if a][:200]
    if not values or not actor_names or tenant_cfg.mode == "disabled":
        return {}
    system = (
        "You match free-text person references from a spreadsheet to the "
        "closest name in a resource pool. Reply ONLY strict JSON "
        '{"<raw value>": "<exact pool name or null>"}. Use null unless '
        "you are confident it is the same person. No prose."
    )
    try:
        res = await generate_for_tenant(
            envolver_no_confiable(
                json.dumps(
                    {"values": values, "pool": actor_names}, ensure_ascii=False
                ),
                origen="nombres del archivo subido y del catálogo de recursos",
            ),
            system=build_system_prompt(system, None),
            tenant_ai_mode=tenant_cfg.mode,
            byo_config=tenant_cfg.byo,
            tenant_id=tenant_id,
            json_mode=True,
        )
        parsed = _extract_json(res.text)
    except Exception as exc:
        log.info("ai_match_resources fallo: %s", exc)
        return {}
    if not isinstance(parsed, dict):
        return {}
    pool = set(actor_names)
    return {
        raw: name
        for raw, name in parsed.items()
        if raw in values and isinstance(name, str) and name in pool
    }


_STRUCTURE_SYSTEM = (
    "You convert a raw spreadsheet (list of rows, possibly messy: section "
    "titles, indentation instead of WBS codes, headers anywhere or absent) "
    "into a project plan. Reply ONLY strict JSON: a list of tasks "
    '[{"wbs_code": "1.2" or null, "name": str (required), '
    '"start_date": "YYYY-MM-DD" or null, "end_date": "YYYY-MM-DD" or null, '
    '"progress": 0-100 or null, '
    '"status": "not_started|in_progress|completed|on_hold" or null, '
    '"is_milestone": bool}]. Rules: skip header/empty/total rows; if rows '
    "are grouped under section titles, make the section a parent task and "
    "derive hierarchical WBS codes (1, 1.1, 1.2, 2, ...); preserve original "
    "WBS codes when present (as text, e.g. '1.30' stays '1.30'); percentages "
    "may come as 0-1 fractions (0.45 == 45). No prose, JSON only."
)


async def ai_propose_structure(
    rows: list[list[str | None]],
    *,
    tenant_cfg: TenantAIConfig,
    tenant_id: str | None = None,
) -> list[ParsedTask]:
    """Nivel 3 — propone un plan completo desde filas crudas. Valida y
    coerciona server-side cada tarea propuesta; devuelve [] si el LLM
    falla o no produce nada usable."""
    if tenant_cfg.mode == "disabled" or not rows:
        return []
    trimmed = [
        [None if c is None else str(c)[:200] for c in r[:MAX_STRUCTURE_COLS]]
        for r in rows[:MAX_STRUCTURE_ROWS]
    ]
    try:
        res = await generate_for_tenant(
            envolver_no_confiable(
                json.dumps({"rows": trimmed}, ensure_ascii=False),
                origen="filas crudas del archivo subido por el usuario",
            ),
            system=build_system_prompt(_STRUCTURE_SYSTEM, None),
            tenant_ai_mode=tenant_cfg.mode,
            byo_config=tenant_cfg.byo,
            tenant_id=tenant_id,
            json_mode=True,
        )
        parsed = _extract_json(res.text)
    except Exception as exc:
        log.info("ai_propose_structure fallo: %s", exc)
        return []
    if isinstance(parsed, dict):
        # Algunos modelos envuelven la lista: {"tasks": [...]}.
        parsed = parsed.get("tasks")
    if not isinstance(parsed, list):
        return []

    out: list[ParsedTask] = []
    for i, item in enumerate(parsed[:MAX_STRUCTURE_ROWS], start=2):
        if not isinstance(item, dict):
            continue
        name = str(item.get("name") or "").strip()
        if not name:
            continue
        wbs_raw = item.get("wbs_code")
        wbs_code = str(wbs_raw).strip() if wbs_raw not in (None, "") else None
        prog_raw = item.get("progress")
        try:
            prog = float(prog_raw) if prog_raw is not None else 0.0
        except (TypeError, ValueError):
            prog = 0.0
        if 0 < prog <= 1:
            prog *= 100
        out.append(
            ParsedTask(
                row_number=i,
                name=name[:300],
                wbs_code=wbs_code,
                start_date=_coerce_date(item.get("start_date")),
                end_date=_coerce_date(item.get("end_date")),
                progress=max(0, min(100, round(prog))),
                status=_coerce_status(item.get("status")),
                is_milestone=bool(item.get("is_milestone")),
            )
        )
    return out
