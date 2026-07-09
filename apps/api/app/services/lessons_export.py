"""Lessons Learned export — ENH-187.

Genera un Excel de **1 hoja "Lecciones"** con columnas en español. Mismo
estilo que `change_export.py` (ENH-186) / `raid_export.py` (ENH-152):
header bold con fondo del DS, freeze pane en fila 2 y autosize
best-effort. Las filas llegan ya formateadas (responsable resuelto a
texto en el endpoint), siguiendo el patrón de `organigrama_export.py`.
"""
from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
from typing import Any

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

LESSON_HEADERS: list[str] = [
    "Folio",
    "Lección",
    "Descripción",
    "Categoría",
    "Fase",
    "Responsable",
    "Recomendación",
    "Tags",
    "Estado",
]

# Consistente con LESSON_CATEGORY_LABEL en apps/web/lib/api/modules.ts.
LESSON_CATEGORY_ES: dict[str, str] = {
    "success": "Éxito",
    "improvement": "Mejora",
    "error": "Error",
}

# Consistente con LESSON_PHASE_LABEL en apps/web/lib/api/modules.ts (valores
# libres en DB; el modal de creación sólo ofrece estos 4).
LESSON_PHASE_ES: dict[str, str] = {
    "planning": "Planificación",
    "execution": "Ejecución",
    "support": "Soporte",
    "closed": "Cierre",
}

LESSON_STATUS_ES: dict[str, str] = {
    "published": "Publicada",
    "draft": "Borrador",
    "archived": "Archivada",
}


def build_lesson_rows(
    lessons: list[Any],
    actor_names: dict[str, str],
) -> list[list[Any]]:
    """Filas de la hoja Lecciones. `actor_names` resuelve owner_actor_id
    (UUID de Actor) a nombre para mostrar texto en vez de UUID."""
    rows: list[list[Any]] = []
    for l in lessons:
        rows.append(
            [
                l.folio,
                l.title,
                l.description or "",
                LESSON_CATEGORY_ES.get(l.category, l.category) if l.category else "",
                LESSON_PHASE_ES.get(l.phase, l.phase) if l.phase else "",
                actor_names.get(str(l.owner_actor_id), "") if l.owner_actor_id else "",
                l.recommendation or "",
                l.tags or [],
                LESSON_STATUS_ES.get(l.status, l.status),
            ]
        )
    return rows


def _cell_value(value: Any) -> Any:
    """Convierte tipos no-Excel-friendly. tz-aware datetimes → naive (Excel
    no acepta tz). Listas/dicts → string."""
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.replace(tzinfo=None) if value.tzinfo else value
    if isinstance(value, date):
        return value
    if isinstance(value, (list, dict)):
        return ", ".join(str(x) for x in value) if isinstance(value, list) else str(value)
    return value


def _autosize(ws, n_cols: int) -> None:
    """Best-effort: ancho = max longitud de la columna, cap a 60."""
    from openpyxl.utils import get_column_letter

    for idx in range(1, n_cols + 1):
        col_letter = get_column_letter(idx)
        max_len = 0
        for cell in ws[col_letter]:
            try:
                v = "" if cell.value is None else str(cell.value)
                if len(v) > max_len:
                    max_len = len(v)
            except Exception:
                continue
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)


def _write_sheet(wb, title: str, headers: list[str], rows: list[list[Any]]) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill

    ws = wb.create_sheet(title=title)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_align = Alignment(vertical="center")

    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    for row in rows:
        ws.append([_cell_value(v) for v in row])

    ws.freeze_panes = "A2"
    _autosize(ws, len(headers))


def export_lessons_xlsx(rows: list[list[Any]]) -> bytes:
    """Devuelve bytes XLSX con 1 hoja "Lecciones". Si no hay filas, incluye
    igual el header."""
    from openpyxl import Workbook

    wb = Workbook()
    default_ws = wb.active
    if default_ws is not None:
        wb.remove(default_ws)
    _write_sheet(wb, "Lecciones", LESSON_HEADERS, rows)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
