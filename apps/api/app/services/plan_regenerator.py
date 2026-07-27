"""Plan regenerator (ENH-080 · ENH-193).

Regenera el archivo del Plan on-demand a partir de las tareas en DB.
DB es la fuente de verdad; el archivo se reconstruye preservando el formato
origen detectado al subir (`source_format` ∈ {mpp, xlsx, csv, template}).

ENH-193: el contrato de columnas es el MISMO de la plantilla V1 del
frontend (`apps/web/lib/plan-template.ts`) y del export de la página del
Plan — 15 columnas — para que el round-trip download → editar → re-subir
funcione sin mapeo manual. El orden de filas es el orden real del plan
(position manual → WBS natural), no outline-first.

- xlsx / template → openpyxl con plantilla mínima (headers + filas).
- csv             → flat export sin fórmulas.
- mpp             → MS Project binario es write-only via MPXJ writer (Java);
  no soportado en este sprint. Fallback: regenera xlsx y devuelve header
  `X-Plan-Format-Fallback: xlsx-mpp-not-supported` para que la UI muestre
  warning. Limitación documentada en US-310 CA3.
"""
from __future__ import annotations

import csv
from dataclasses import dataclass, field
from io import BytesIO, StringIO

from app.models.task import Task
from app.services.plan_metadata import wbs_sort_key

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
CSV_MIME = "text/csv"

# ENH-193: espeja `COLUMNS` de apps/web/lib/plan-template.ts (V1).
PLAN_HEADERS = [
    "WBS",
    "Tarea",
    "Outline Level",
    "Inicio",
    "Fin",
    "Duración (días)",
    "Avance (%)",
    "Estado",
    "Área Responsable",
    "Responsable",
    "Criticidad",
    "Es hito",
    "Hito Relacionado",
    "Predecessors",
    "Successors",
]


@dataclass
class PlanExportContext:
    """Lookups para resolver FKs a texto legible en el export.

    Todas opcionales: sin contexto el export degrada a celdas vacías
    (nunca IDs crudos).
    """

    area_names: dict[str, str] = field(default_factory=dict)  # area_id → nombre
    actor_names: dict[str, str] = field(default_factory=dict)  # actor_id → nombre
    milestone_wbs: dict[str, str] = field(default_factory=dict)  # task_id → wbs


def plan_order(tasks: list[Task]) -> list[Task]:
    """Orden real del plan: `position` manual manda (US-176), después
    WBS natural (BUG-049). Igual que `GET /tasks`."""
    return sorted(
        tasks,
        key=lambda t: (
            t.position is None,
            t.position if t.position is not None else 0,
            wbs_sort_key(t.wbs),
        ),
    )


def _row(task: Task, ctx: PlanExportContext) -> list[object]:
    return [
        task.wbs or "",
        task.name or "",
        task.outline_level if task.outline_level is not None else "",
        task.start_date.isoformat() if task.start_date else "",
        task.end_date.isoformat() if task.end_date else "",
        task.duration_days if task.duration_days is not None else "",
        task.progress if task.progress is not None else 0,
        task.status or "",
        ctx.area_names.get(str(task.area_id), "") if task.area_id else "",
        (
            ctx.actor_names.get(str(task.assignee_actor_id), "")
            if task.assignee_actor_id
            else ""
        ),
        # Plantilla V1: Criticidad booleana Sí/No (ENH-134).
        "Sí" if getattr(task, "is_critical", False) else "No",
        "Sí" if task.is_milestone else "No",
        (
            ctx.milestone_wbs.get(str(task.related_milestone_id), "")
            if task.related_milestone_id
            else ""
        ),
        ", ".join(task.predecessors or []),
        ", ".join(task.successors or []),
    ]


def regenerate_xlsx(tasks: list[Task], ctx: PlanExportContext | None = None) -> bytes:
    """xlsx: 1 sheet "Plan" con headers + filas en el orden real del
    plan. Columnas = plantilla V1, así el round-trip download →
    re-upload preserva todos los datos (incl. Estado, BUG-088: WBS en
    formato texto)."""
    from openpyxl import Workbook

    ctx = ctx or PlanExportContext()
    wb = Workbook()
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet("Plan")
    else:
        ws.title = "Plan"
    ws.append(PLAN_HEADERS)
    for t in plan_order(tasks):
        ws.append(_row(t, ctx))
    # BUG-088: WBS como texto para que ediciones en Excel no lo
    # conviertan a número (1.30 → 1.3).
    for row_idx in range(2, ws.max_row + 1):
        ws.cell(row=row_idx, column=1).number_format = "@"
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def regenerate_csv(tasks: list[Task], ctx: PlanExportContext | None = None) -> bytes:
    """csv flat (UTF-8 sin BOM). Misma columna order que xlsx."""
    ctx = ctx or PlanExportContext()
    buf = StringIO()
    writer = csv.writer(buf)
    writer.writerow(PLAN_HEADERS)
    for t in plan_order(tasks):
        writer.writerow(_row(t, ctx))
    return buf.getvalue().encode("utf-8")


def regenerate_for_format(
    fmt: str, tasks: list[Task], ctx: PlanExportContext | None = None
) -> tuple[bytes, str, str, bool]:
    """Devuelve `(bytes, mime, ext, fallback_used)`.

    `fallback_used=True` cuando el formato pedido era `mpp` y se cayó a xlsx.
    """
    fmt = (fmt or "xlsx").lower()
    if fmt == "csv":
        return regenerate_csv(tasks, ctx), CSV_MIME, "csv", False
    if fmt == "mpp":
        # MPP write requiere MPXJ Pro / paquete comercial; no viable en
        # plataforma. Fallback a xlsx con warning header (US-310 CA3).
        return regenerate_xlsx(tasks, ctx), XLSX_MIME, "xlsx", True
    # xlsx | template | desconocido → xlsx default.
    return regenerate_xlsx(tasks, ctx), XLSX_MIME, "xlsx", False
